#!/usr/bin/env python3
import argparse
import json
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


def _norm(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _is_blank(value):
    return _norm(value) == ""


def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dict):
        pairs = [f"{k}={_cell_value(v)}" for k, v in value.items()]
        return "; ".join(pairs)
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(_cell_value(v)) for v in value)
    return value


def _display_text(value):
    return str(_cell_value(value)).strip()


def _equals_display(new_raw, old_value):
    return str(_norm(new_raw)) == _display_text(old_value)


def _parse_bool(value):
    v = _norm(value)
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.lower()
        if s in {"true", "t", "yes", "y", "1"}:
            return True
        if s in {"false", "f", "no", "n", "0", ""}:
            return False
    return bool(v)


def _parse_number(value):
    v = _norm(value)
    if v == "":
        return ""
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        try:
            if "." in v:
                return float(v)
            return int(v)
        except ValueError:
            return v
    return v


def _split_csv(value):
    v = _norm(value)
    if v == "":
        return []
    if isinstance(v, str):
        return [x.strip() for x in v.split(",") if x.strip()]
    return [str(v)]


def _parse_sensitive_fields(value):
    v = _norm(value)
    if v == "":
        return {}
    out = {}
    if isinstance(v, str):
        parts = [p.strip() for p in v.split(";") if p.strip()]
        for p in parts:
            if ":" in p:
                k, val = p.split(":", 1)
                out[k.strip()] = val.strip()
            else:
                out[p] = ""
    return out


def _parse_json_text(value):
    v = _norm(value)
    if v == "":
        return ""
    if isinstance(v, (dict, list)):
        return v
    try:
        return json.loads(str(v))
    except Exception:
        return v


def _coerce_like(old_value, new_raw):
    if old_value is None:
        return _norm(new_raw)
    if isinstance(old_value, bool):
        return _parse_bool(new_raw)
    if isinstance(old_value, int) and not isinstance(old_value, bool):
        v = _parse_number(new_raw)
        return int(v) if isinstance(v, (int, float)) and str(v) != "" else old_value
    if isinstance(old_value, float):
        v = _parse_number(new_raw)
        return float(v) if isinstance(v, (int, float)) and str(v) != "" else old_value
    if isinstance(old_value, list):
        return _split_csv(new_raw)
    if isinstance(old_value, dict):
        parsed = _parse_json_text(new_raw)
        return parsed if isinstance(parsed, dict) else old_value
    return _norm(new_raw)


def _set_if_changed(obj, key, new_raw, parser=None):
    old = obj.get(key)
    if _equals_display(new_raw, old):
        return
    if parser is None:
        obj[key] = _coerce_like(old, new_raw)
    else:
        obj[key] = parser(new_raw)


def _sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if row is None:
            continue
        item = {}
        has_data = False
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            item[header] = val
            if not _is_blank(val):
                has_data = True
        if has_data:
            out.append(item)
    return out


def _read_multi_sections(ws):
    rows = list(ws.iter_rows(values_only=True))
    sections = {}
    i = 0
    while i < len(rows):
        row = rows[i] or ()
        section_name = row[0] if row else None
        if _is_blank(section_name):
            i += 1
            continue

        if i + 1 >= len(rows):
            break
        header_row = rows[i + 1] or ()
        headers = [str(h).strip() for h in header_row if not _is_blank(h)]
        i += 2

        data = []
        while i < len(rows):
            cur = rows[i] or ()
            if all(_is_blank(c) for c in cur):
                i += 1
                break
            item = {}
            has_data = False
            for idx, header in enumerate(headers):
                val = cur[idx] if idx < len(cur) else None
                item[header] = val
                if not _is_blank(val):
                    has_data = True
            if has_data:
                data.append(item)
            i += 1

        sections[str(section_name)] = data
    return sections


def _read_roundtrip_payload(wb):
    if "__rt_payload" not in wb.sheetnames:
        raise ValueError("Workbook missing hidden __rt_payload sheet. Re-export with updated json_to_excel.py.")

    rows = _sheet_rows(wb["__rt_payload"])
    chunks = []
    for row in rows:
        idx = row.get("chunk_index")
        chunk = row.get("payload_chunk")
        if _is_blank(chunk):
            continue
        try:
            idx_int = int(idx)
        except Exception:
            idx_int = len(chunks) + 1
        chunks.append((idx_int, str(chunk)))

    if not chunks:
        raise ValueError("__rt_payload is empty.")

    chunks.sort(key=lambda x: x[0])
    payload_json = "".join(chunk for _, chunk in chunks)
    return json.loads(payload_json)


def _derive_database(connection, metadata):
    db = connection.get("database")
    if db:
        return db
    db_url = metadata.get("database_url") or ""
    if not db_url:
        return ""
    parsed = urlparse(db_url if "://" in db_url else f"dummy://{db_url}")
    if parsed.path and parsed.path != "/":
        return parsed.path.lstrip("/")
    q = parse_qs(parsed.query or "")
    service = q.get("service_name", [""])[0]
    return service or ""


def _tables_index(payload):
    idx = {}
    for table in payload.get("tables", []) or []:
        key = (str(table.get("schema", "")).strip(), str(table.get("table", "")).strip())
        idx[key] = table
    return idx


def _columns_index(table):
    idx = {}
    for col in table.get("columns", []) or []:
        idx[str(col.get("name", "")).strip()] = col
    return idx


def _ensure_table(payload, tindex, schema, table_name):
    key = (schema, table_name)
    if key in tindex:
        return tindex[key]
    table = {
        "schema": schema,
        "table": table_name,
        "columns": [],
        "foreign_keys": [],
        "join_candidates": [],
        "sample_data": {},
    }
    payload.setdefault("tables", []).append(table)
    tindex[key] = table
    return table


def _ensure_column(table, column_name):
    cidx = _columns_index(table)
    if column_name in cidx:
        return cidx[column_name]
    col = {"name": column_name}
    table.setdefault("columns", []).append(col)
    return col


def _apply_summary_sheet(wb, payload):
    if "Summary" not in wb.sheetnames:
        return
    sections = _read_multi_sections(wb["Summary"])
    connection = payload.setdefault("connection", {})
    metadata = payload.setdefault("metadata", {})
    dqs = payload.setdefault("data_quality_summary", {})

    overview = sections.get("Overview", [])
    for row in overview:
        metric = str(_norm(row.get("metric")))
        value = row.get("value")
        if metric == "source_name":
            display = connection.get("source_name") or connection.get("host") or metadata.get("database_url") or ""
            if not _equals_display(value, display):
                connection["source_name"] = _norm(value)
        elif metric == "source_type":
            display = connection.get("source_type") or connection.get("driver") or ""
            if not _equals_display(value, display):
                connection["source_type"] = _norm(value)
        elif metric == "database":
            display = _derive_database(connection, metadata)
            if not _equals_display(value, display):
                connection["database"] = _norm(value)
        elif metric == "schema":
            display = connection.get("schema") or metadata.get("schema_filter") or ""
            if not _equals_display(value, display):
                connection["schema"] = _norm(value)
        elif metric == "port":
            _set_if_changed(connection, "port", value, parser=lambda v: _coerce_like(connection.get("port"), v))
        elif metric == "database_timezone":
            _set_if_changed(connection, "timezone", value)
        elif metric.startswith("data_quality_"):
            key = metric[len("data_quality_") :]
            old = dqs.get(key)
            if not _equals_display(value, old):
                dqs[key] = _coerce_like(old, value)

    by_check = dqs.setdefault("by_check", {})
    for row in sections.get("DataQualityByCheck", []):
        check = str(_norm(row.get("check")))
        if not check:
            continue
        value = row.get("count")
        old = by_check.get(check)
        if not _equals_display(value, old):
            by_check[check] = _coerce_like(old, value)

    constraints = dqs.setdefault("constraints_found", {})
    for row in sections.get("DataQualityConstraints", []):
        key = str(_norm(row.get("constraint")))
        if not key:
            continue
        value = row.get("count")
        old = constraints.get(key)
        if not _equals_display(value, old):
            constraints[key] = _coerce_like(old, value)

    for row in sections.get("DataQualityDetails", []):
        group = str(_norm(row.get("metric_group")))
        item = str(_norm(row.get("item")))
        if not group.startswith("data_quality_") or not item:
            continue
        key = group[len("data_quality_") :]
        grp = dqs.setdefault(key, {})
        old = grp.get(item)
        value = row.get("value")
        if not _equals_display(value, old):
            grp[item] = _coerce_like(old, value)


def _row_has_any(row, keys):
    return any(not _is_blank(row.get(k)) for k in keys)


def _apply_source_context_sheet(wb, payload):
    if "SourceContextManual" not in wb.sheetnames:
        return
    sections = _read_multi_sections(wb["SourceContextManual"])
    sctx = payload.setdefault("source_system_context", {})

    contacts = []
    for row in sections.get("ContactsManual", []):
        if _row_has_any(row, ["contact_name", "role", "email", "phone", "notes"]):
            contacts.append(
                {
                    "name": _norm(row.get("contact_name")),
                    "role": _norm(row.get("role")),
                    "email": _norm(row.get("email")),
                    "phone": _norm(row.get("phone")),
                    "notes": _norm(row.get("notes")),
                }
            )
    if contacts:
        sctx["contacts"] = contacts

    delete_rows = []
    for row in sections.get("DeleteManagementManual", []):
        if _row_has_any(row, ["table_name", "delete_strategy", "instruction", "notes"]):
            delete_rows.append(
                {
                    "table_name": _norm(row.get("table_name")),
                    "delete_strategy": _norm(row.get("delete_strategy")),
                    "instruction": _norm(row.get("instruction")),
                    "notes": _norm(row.get("notes")),
                }
            )
    if delete_rows:
        sctx["delete_management_instruction"] = delete_rows[0].get("instruction", "")
        sctx["delete_management_manual_rows"] = delete_rows

    restrictions = []
    for row in sections.get("RestrictionsManual", []):
        if _row_has_any(row, ["restriction_type", "scope", "details", "owner"]):
            restrictions.append(
                {
                    "type": _norm(row.get("restriction_type")),
                    "scope": _norm(row.get("scope")),
                    "details": _norm(row.get("details")),
                    "owner": _norm(row.get("owner")),
                }
            )
    if restrictions:
        old = sctx.get("restrictions")
        if isinstance(old, dict) and len(restrictions) == 1:
            sctx["restrictions"] = restrictions[0]
        else:
            sctx["restrictions"] = restrictions

    late_rows = []
    for row in sections.get("LateArrivingDataManual", []):
        if _row_has_any(row, ["table_name", "business_date_column", "system_ts_column", "lookback_days", "policy_notes"]):
            late_rows.append(
                {
                    "table_name": _norm(row.get("table_name")),
                    "business_date_column": _norm(row.get("business_date_column")),
                    "system_ts_column": _norm(row.get("system_ts_column")),
                    "lookback_days": _norm(row.get("lookback_days")),
                    "policy_notes": _norm(row.get("policy_notes")),
                }
            )
    if late_rows:
        sctx["late_arriving_data_manual"] = late_rows[0].get("policy_notes", "")
        sctx["late_arriving_data_manual_rows"] = late_rows

    vol_rows = []
    for row in sections.get("VolumeSizeProjectionManual", []):
        if _row_has_any(row, ["entity_scope", "projection_horizon_months", "growth_assumption_pct", "basis", "notes"]):
            vol_rows.append(
                {
                    "entity_scope": _norm(row.get("entity_scope")),
                    "projection_horizon_months": _norm(row.get("projection_horizon_months")),
                    "growth_assumption_pct": _norm(row.get("growth_assumption_pct")),
                    "basis": _norm(row.get("basis")),
                    "notes": _norm(row.get("notes")),
                }
            )
    if vol_rows:
        sctx["volume_size_projection_manual"] = vol_rows[0].get("notes", "")
        sctx["volume_size_projection_manual_rows"] = vol_rows

    fc_rows = []
    for row in sections.get("FieldContextManual", []):
        if _row_has_any(row, ["table_name", "column_name", "business_context", "transformation_notes", "owner"]):
            fc_rows.append(
                {
                    "table_name": _norm(row.get("table_name")),
                    "column_name": _norm(row.get("column_name")),
                    "business_context": _norm(row.get("business_context")),
                    "transformation_notes": _norm(row.get("transformation_notes")),
                    "owner": _norm(row.get("owner")),
                }
            )
    if fc_rows:
        sctx["field_context_manual"] = fc_rows[0].get("business_context", "")
        sctx["field_context_manual_rows"] = fc_rows


def _apply_tables_sheet(wb, payload, tindex):
    if "Tables" not in wb.sheetnames:
        return
    for row in _sheet_rows(wb["Tables"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        if not schema or not table_name:
            continue
        table = _ensure_table(payload, tindex, schema, table_name)

        _set_if_changed(table, "row_count", row.get("row_count"), parser=lambda v: _coerce_like(table.get("row_count"), v))
        _set_if_changed(table, "has_primary_key", row.get("has_primary_key"), parser=_parse_bool)
        _set_if_changed(table, "primary_keys", row.get("primary_keys"), parser=_split_csv)
        _set_if_changed(table, "has_foreign_keys", row.get("has_foreign_keys"), parser=_parse_bool)
        _set_if_changed(table, "incremental_columns", row.get("incremental_columns"), parser=_split_csv)
        _set_if_changed(table, "partition_columns", row.get("partition_columns"), parser=_split_csv)
        _set_if_changed(table, "partition_columns_candidates", row.get("partition_columns_candidates"), parser=_split_csv)
        _set_if_changed(table, "sensitive_fields", row.get("sensitive_fields"), parser=_parse_sensitive_fields)
        _set_if_changed(table, "table_description", row.get("table_description"))
        _set_if_changed(table, "cdc_enabled", row.get("cdc_enabled"), parser=_parse_bool)


def _apply_columns_row(col, row):
    _set_if_changed(col, "type", row.get("data_type"))
    _set_if_changed(col, "nullable", row.get("nullable"), parser=_parse_bool)
    _set_if_changed(col, "is_incremental", row.get("is_incremental"), parser=_parse_bool)
    _set_if_changed(col, "cardinality", row.get("cardinality"), parser=lambda v: _coerce_like(col.get("cardinality"), v))
    _set_if_changed(col, "null_count", row.get("null_count"), parser=lambda v: _coerce_like(col.get("null_count"), v))
    _set_if_changed(col, "data_category", row.get("data_category"))
    _set_if_changed(col, "semantic_class", row.get("semantic_class"))
    _set_if_changed(col, "description", row.get("description"))

    existing_uc = col.get("unit_context") if isinstance(col.get("unit_context"), dict) else {}
    display_unit = col.get("unit")
    if _is_blank(display_unit):
        display_unit = existing_uc.get("detected_unit")
    display_unit_source = col.get("unit_source")
    if _is_blank(display_unit_source):
        display_unit_source = existing_uc.get("detection_source")
    if not _equals_display(row.get("unit"), display_unit):
        col["unit"] = _coerce_like(col.get("unit"), row.get("unit"))
    if not _equals_display(row.get("unit_source"), display_unit_source):
        col["unit_source"] = _coerce_like(col.get("unit_source"), row.get("unit_source"))

    has_existing_conv = isinstance(existing_uc.get("conversion"), dict)
    existing_conv = existing_uc.get("conversion") if has_existing_conv else {}
    needs_conv = any(
        [
            not _equals_display(row.get("factor_to_canonical"), existing_conv.get("factor_to_canonical")),
            not _equals_display(row.get("offset_to_canonical"), existing_conv.get("offset_to_canonical")),
            not _equals_display(row.get("conversion_formula"), existing_conv.get("formula")),
        ]
    )
    needs_uc = any(
        [
            not _equals_display(row.get("unit"), existing_uc.get("detected_unit")),
            not _equals_display(row.get("unit_source"), existing_uc.get("detection_source")),
            not _equals_display(row.get("canonical_unit"), existing_uc.get("canonical_unit")),
            not _equals_display(row.get("unit_system"), existing_uc.get("unit_system")),
            not _equals_display(row.get("unit_confidence"), existing_uc.get("detection_confidence")),
            not _equals_display(row.get("unit_notes"), existing_uc.get("notes")),
            needs_conv,
        ]
    )
    if needs_uc or isinstance(col.get("unit_context"), dict):
        unit_context = col.get("unit_context")
        if not isinstance(unit_context, dict):
            unit_context = {}
            col["unit_context"] = unit_context
        conversion = unit_context.get("conversion")

        _set_if_changed(unit_context, "detected_unit", row.get("unit"))
        _set_if_changed(unit_context, "detection_source", row.get("unit_source"))
        _set_if_changed(unit_context, "canonical_unit", row.get("canonical_unit"))
        _set_if_changed(unit_context, "unit_system", row.get("unit_system"))
        _set_if_changed(unit_context, "detection_confidence", row.get("unit_confidence"), parser=lambda v: _coerce_like(unit_context.get("detection_confidence"), v))
        _set_if_changed(unit_context, "notes", row.get("unit_notes"))
        if needs_conv or has_existing_conv:
            if not isinstance(conversion, dict):
                conversion = {}
                unit_context["conversion"] = conversion
            _set_if_changed(conversion, "factor_to_canonical", row.get("factor_to_canonical"), parser=lambda v: _coerce_like(conversion.get("factor_to_canonical"), v))
            _set_if_changed(conversion, "offset_to_canonical", row.get("offset_to_canonical"), parser=lambda v: _coerce_like(conversion.get("offset_to_canonical"), v))
            _set_if_changed(conversion, "formula", row.get("conversion_formula"))

    existing_dr = col.get("data_range") if isinstance(col.get("data_range"), dict) else {}
    needs_dr = any(
        [
            not _equals_display(row.get("range_min"), existing_dr.get("min")),
            not _equals_display(row.get("range_max"), existing_dr.get("max")),
        ]
    )
    if needs_dr or isinstance(col.get("data_range"), dict):
        data_range = col.get("data_range")
        if not isinstance(data_range, dict):
            data_range = {}
            col["data_range"] = data_range
        _set_if_changed(data_range, "min", row.get("range_min"), parser=lambda v: _coerce_like(data_range.get("min"), v))
        _set_if_changed(data_range, "max", row.get("range_max"), parser=lambda v: _coerce_like(data_range.get("max"), v))


def _apply_columns_sheet(wb, payload, tindex):
    if "Columns" not in wb.sheetnames:
        return

    for row in _sheet_rows(wb["Columns"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        column_name = str(_norm(row.get("column_name")))
        if not schema or not table_name or not column_name:
            continue

        table = _ensure_table(payload, tindex, schema, table_name)
        col = _ensure_column(table, column_name)
        _apply_columns_row(col, row)


def _apply_units_sheet(wb, payload, tindex):
    if "Units" not in wb.sheetnames:
        return

    for row in _sheet_rows(wb["Units"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        column_name = str(_norm(row.get("column_name")))
        if not schema or not table_name or not column_name:
            continue
        table = _ensure_table(payload, tindex, schema, table_name)
        col = _ensure_column(table, column_name)
        _apply_columns_row(
            col,
            {
                "unit": row.get("unit"),
                "unit_source": row.get("unit_source"),
                "canonical_unit": row.get("canonical_unit"),
                "unit_system": row.get("unit_system"),
                "unit_confidence": row.get("unit_confidence"),
                "unit_notes": row.get("unit_notes"),
                "factor_to_canonical": row.get("factor_to_canonical"),
                "offset_to_canonical": row.get("offset_to_canonical"),
                "conversion_formula": row.get("conversion_formula"),
                "data_type": col.get("type", ""),
                "nullable": col.get("nullable", ""),
                "is_incremental": col.get("is_incremental", ""),
                "cardinality": col.get("cardinality", ""),
                "null_count": col.get("null_count", ""),
                "data_category": col.get("data_category", ""),
                "semantic_class": col.get("semantic_class", ""),
                "description": col.get("description", ""),
                "range_min": (col.get("data_range") or {}).get("min", ""),
                "range_max": (col.get("data_range") or {}).get("max", ""),
            },
        )


def _apply_foreign_keys(wb, tindex):
    if "ForeignKeys" not in wb.sheetnames:
        return

    grouped = {}
    for row in _sheet_rows(wb["ForeignKeys"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        if not schema or not table_name:
            continue
        tkey = (schema, table_name)
        if tkey not in tindex:
            continue
        col = _norm(row.get("column_name"))
        ref = _norm(row.get("references"))
        if not col and not ref:
            continue
        grouped.setdefault(tkey, []).append({"column": col, "references": ref})

    for tkey, rows in grouped.items():
        table = tindex[tkey]
        existing = table.setdefault("foreign_keys", [])
        emap = {(str(x.get("column", "")), str(x.get("references", ""))): x for x in existing if isinstance(x, dict)}
        for item in rows:
            k = (item["column"], item["references"])
            if k not in emap:
                existing.append(item)
        table["has_foreign_keys"] = len(existing) > 0


def _apply_join_candidates(wb, tindex):
    if "JoinCandidates" not in wb.sheetnames:
        return

    grouped = {}
    for row in _sheet_rows(wb["JoinCandidates"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        if not schema or not table_name:
            continue
        tkey = (schema, table_name)
        if tkey not in tindex:
            continue
        candidate = {
            "column": _norm(row.get("column_name")),
            "target_table": _norm(row.get("target_table")),
            "target_column": _norm(row.get("target_column")),
            "confidence": _parse_number(row.get("confidence")),
        }
        if not candidate["column"] and not candidate["target_table"] and not candidate["target_column"]:
            continue
        grouped.setdefault(tkey, []).append(candidate)

    for tkey, rows in grouped.items():
        table = tindex[tkey]
        existing = table.setdefault("join_candidates", [])
        emap = {}
        for idx, item in enumerate(existing):
            if isinstance(item, dict):
                emap[(str(item.get("column", "")), str(item.get("target_table", "")), str(item.get("target_column", "")))] = idx
        for row in rows:
            k = (row["column"], row["target_table"], row["target_column"])
            if k in emap:
                old = existing[emap[k]]
                if isinstance(old, dict):
                    _set_if_changed(old, "confidence", row.get("confidence"), parser=lambda v: _coerce_like(old.get("confidence"), v))
            else:
                existing.append(row)


def _apply_sample_data(wb, tindex):
    if "SampleData" not in wb.sheetnames:
        return

    for row in _sheet_rows(wb["SampleData"]):
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        sample_col = str(_norm(row.get("sample_column")))
        if not schema or not table_name or not sample_col:
            continue
        tkey = (schema, table_name)
        if tkey not in tindex:
            continue
        table = tindex[tkey]
        sample_data = table.setdefault("sample_data", {})
        values = sample_data.setdefault(sample_col, [])

        idx_raw = _parse_number(row.get("sample_index"))
        try:
            idx = int(idx_raw) - 1
        except Exception:
            idx = len(values)

        while idx >= len(values):
            values.append("")
        if idx < 0:
            idx = 0

        old_val = values[idx] if idx < len(values) else ""
        new_val = row.get("sample_value")
        if not _equals_display(new_val, old_val):
            values[idx] = _coerce_like(old_val, new_val)


def _coerce_findings_field(old_value, new_raw, key):
    if key in {"has_audit_trail"}:
        return _parse_bool(new_raw)
    if key in {"tz_aware_count", "tz_naive_count", "recommended_lookback_days"}:
        return _parse_number(new_raw)
    if key in {"distinct_values", "suggested_domain", "sample_values", "distinct_timezones"}:
        if isinstance(old_value, list):
            return _split_csv(new_raw)
        return _coerce_like(old_value, new_raw)
    return _coerce_like(old_value, new_raw)


def _apply_data_quality_findings(wb, tindex):
    if "DataQualityFindings" not in wb.sheetnames:
        return

    rows = _sheet_rows(wb["DataQualityFindings"])
    for row in rows:
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        if not schema or not table_name:
            continue
        tkey = (schema, table_name)
        if tkey not in tindex:
            continue

        table = tindex[tkey]
        dq = table.setdefault("data_quality", {})
        findings = dq.setdefault("findings", [])

        try:
            find_idx = int(_parse_number(row.get("finding_index"))) - 1
        except Exception:
            find_idx = len(findings)
        if find_idx < 0:
            find_idx = 0

        while find_idx >= len(findings):
            findings.append({})

        finding = findings[find_idx]
        if not isinstance(finding, dict):
            finding = {"detail": _coerce_like("", finding)}
            findings[find_idx] = finding

        direct_map = {
            "check": "check",
            "severity": "severity",
            "column": "column",
            "detail": "detail",
            "recommendation": "recommendation",
            "distinct_values": "distinct_values",
            "suggested_domain": "suggested_domain",
            "sample_values": "sample_values",
            "cardinality": "cardinality",
            "delete_strategy": "delete_strategy",
            "soft_delete_column": "soft_delete_column",
            "soft_delete_type": "soft_delete_type",
            "has_audit_trail": "has_audit_trail",
            "business_date_column": "business_date_column",
            "system_ts_column": "system_ts_column",
            "server_timezone": "server_timezone",
            "distinct_timezones": "distinct_timezones",
            "tz_aware_count": "tz_aware_count",
            "tz_naive_count": "tz_naive_count",
            "detected_unit": "detected_unit",
            "canonical_unit": "canonical_unit",
        }

        for sheet_key, json_key in direct_map.items():
            new_val = row.get(sheet_key)
            old_val = finding.get(json_key)
            if _equals_display(new_val, old_val):
                continue
            finding[json_key] = _coerce_findings_field(old_val, new_val, json_key)

        tz_columns_raw = row.get("timezone_columns")
        if not _is_blank(tz_columns_raw):
            parsed = _parse_json_text(tz_columns_raw)
            old = finding.get("columns")
            if not _equals_display(tz_columns_raw, json.dumps(old, ensure_ascii=True, separators=(",", ":")) if old not in (None, "") else ""):
                finding["columns"] = parsed

        extra_raw = row.get("extra_json")
        if not _is_blank(extra_raw):
            parsed = _parse_json_text(extra_raw)
            if isinstance(parsed, dict):
                for k, v in parsed.items():
                    old = finding.get(k)
                    if not _equals_display(v, old):
                        finding[k] = v


def _apply_late_arriving_sheet(wb, tindex):
    if "latearivingdata" not in wb.sheetnames:
        return

    rows = _sheet_rows(wb["latearivingdata"])
    for row in rows:
        schema = str(_norm(row.get("schema")))
        table_name = str(_norm(row.get("table_name")))
        if not schema or not table_name:
            continue
        tkey = (schema, table_name)
        if tkey not in tindex:
            continue

        table = tindex[tkey]
        dq = table.setdefault("data_quality", {})
        findings = dq.setdefault("findings", [])

        try:
            idx = int(_parse_number(row.get("finding_index"))) - 1
        except Exception:
            idx = len(findings)
        if idx < 0:
            idx = 0
        while idx >= len(findings):
            findings.append({"check": "late_arriving_data"})

        finding = findings[idx]
        if not isinstance(finding, dict):
            finding = {"check": "late_arriving_data"}
            findings[idx] = finding

        finding.setdefault("check", "late_arriving_data")
        _set_if_changed(finding, "severity", row.get("severity"))
        _set_if_changed(finding, "business_date_column", row.get("business_date_column"))
        _set_if_changed(finding, "system_ts_column", row.get("system_ts_column"))
        _set_if_changed(finding, "recommended_lookback_days", row.get("recommended_lookback_days"), parser=lambda v: _coerce_like(finding.get("recommended_lookback_days"), v))
        _set_if_changed(finding, "detail", row.get("detail"))
        _set_if_changed(finding, "recommendation", row.get("recommendation"))

        lag = finding.setdefault("lag_stats", {})
        _set_if_changed(lag, "total_rows_compared", row.get("lag_total_rows_compared"), parser=lambda v: _coerce_like(lag.get("total_rows_compared"), v))
        _set_if_changed(lag, "min_lag_hours", row.get("lag_min_lag_hours"), parser=lambda v: _coerce_like(lag.get("min_lag_hours"), v))
        _set_if_changed(lag, "avg_lag_hours", row.get("lag_avg_lag_hours"), parser=lambda v: _coerce_like(lag.get("avg_lag_hours"), v))
        _set_if_changed(lag, "p95_lag_hours", row.get("lag_p95_lag_hours"), parser=lambda v: _coerce_like(lag.get("p95_lag_hours"), v))
        _set_if_changed(lag, "max_lag_hours", row.get("lag_max_lag_hours"), parser=lambda v: _coerce_like(lag.get("max_lag_hours"), v))
        _set_if_changed(lag, "max_lag_days", row.get("lag_max_lag_days"), parser=lambda v: _coerce_like(lag.get("max_lag_days"), v))
        _set_if_changed(lag, "rows_late_over_1d", row.get("lag_rows_late_over_1d"), parser=lambda v: _coerce_like(lag.get("rows_late_over_1d"), v))
        _set_if_changed(lag, "rows_late_over_7d", row.get("lag_rows_late_over_7d"), parser=lambda v: _coerce_like(lag.get("rows_late_over_7d"), v))


def apply_all_visible_edits(wb, payload):
    _apply_summary_sheet(wb, payload)
    _apply_source_context_sheet(wb, payload)

    tindex = _tables_index(payload)
    _apply_tables_sheet(wb, payload, tindex)
    _apply_columns_sheet(wb, payload, tindex)
    _apply_units_sheet(wb, payload, tindex)
    _apply_foreign_keys(wb, tindex)
    _apply_join_candidates(wb, tindex)
    _apply_sample_data(wb, tindex)
    _apply_data_quality_findings(wb, tindex)
    _apply_late_arriving_sheet(wb, tindex)


def main():
    parser = argparse.ArgumentParser(description="Convert exported schema Excel workbook back to JSON")
    parser.add_argument("input_xlsx", help="Path to source .xlsx file")
    parser.add_argument("output_json", nargs="?", help="Path to output .json file")
    parser.add_argument(
        "--no-apply-edits",
        action="store_true",
        help="Restore original payload only and ignore visible sheet edits",
    )
    args = parser.parse_args()

    input_path = Path(args.input_xlsx).expanduser().resolve()
    output_path = (
        Path(args.output_json).expanduser().resolve()
        if args.output_json
        else input_path.with_suffix(".json")
    )

    wb = load_workbook(input_path, data_only=True)
    payload = _read_roundtrip_payload(wb)

    if not args.no_apply_edits:
        apply_all_visible_edits(wb, payload)

    output_path.write_text(json.dumps(payload, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote JSON to {output_path}")


if __name__ == "__main__":
    main()
