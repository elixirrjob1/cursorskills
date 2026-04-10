#!/usr/bin/env python3
import argparse
import base64
import json
import logging
import os
import sys
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

_log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# OpenMetadata direct REST helpers (mirrors tools/openmetadata_mcp/server.py)
# ---------------------------------------------------------------------------

def _om_base_url() -> str | None:
    raw = os.getenv("OPENMETADATA_BASE_URL", "").strip().rstrip("/")
    if not raw:
        return None
    return f"{raw}/api" if not raw.endswith("/api") else raw


def _om_login_token() -> str | None:
    jwt = os.getenv("OPENMETADATA_JWT_TOKEN", "").strip()
    if jwt:
        return jwt
    base = _om_base_url()
    email = os.getenv("OPENMETADATA_EMAIL", "").strip()
    password = os.getenv("OPENMETADATA_PASSWORD", "")
    if not base or not email or not password:
        return None
    import requests as _req
    encoded = base64.b64encode(password.encode("utf-8")).decode("ascii")
    for pwd in (encoded, password):
        try:
            resp = _req.post(
                f"{base}/v1/users/login",
                headers={"Accept": "application/json", "Content-Type": "application/json"},
                json={"email": email, "password": pwd},
                timeout=(10, 30),
            )
            resp.raise_for_status()
            data = resp.json() if resp.content else {}
            for key in ("accessToken", "jwtToken", "token", "id_token"):
                val = data.get(key) if isinstance(data, dict) else None
                if isinstance(val, str) and val.strip():
                    return val.strip()
        except Exception:
            continue
    return None


def _om_fetch_glossary_payload() -> dict | None:
    """Fetch glossaries + terms from OpenMetadata REST API. Returns None on any failure."""
    base = _om_base_url()
    if not base:
        return None
    token = _om_login_token()
    if not token:
        _log.warning("OpenMetadata credentials not available; skipping glossary fetch.")
        return None
    import requests as _req
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}",
    }
    try:
        g_resp = _req.get(f"{base}/v1/glossaries", params={"limit": 1000}, headers=headers, timeout=(10, 30))
        g_resp.raise_for_status()
        glossaries = (g_resp.json() or {}).get("data", [])

        t_resp = _req.get(f"{base}/v1/glossaryTerms", params={"limit": 1000}, headers=headers, timeout=(10, 30))
        t_resp.raise_for_status()
        terms = (t_resp.json() or {}).get("data", [])

        _log.info("Fetched %d glossaries and %d terms from OpenMetadata.", len(glossaries), len(terms))
        return {"glossaries": glossaries, "terms": terms}
    except Exception as exc:
        _log.warning("Failed to fetch glossary data from OpenMetadata: %s", exc)
        return None


_GLOSSARY_ROWS = [
    {
        "Field": "glossary_terms",
        "Description": "Comma-separated OpenMetadata glossary term FQNs assigned to the table or column.",
    },
    {
        "Field": "om_class_<ClassificationName>",
        "Description": "One Excel dropdown column per OpenMetadata classification. Each cell stores one allowed tag FQN for that classification, such as PII.Sensitive or Tier.Tier1.",
    },
    {
        "Field": "semantic_class",
        "Description": "Specific meaning inferred for the column, such as email, timestamp, given_name, or customer_identifier.",
    },
    {
        "Field": "concept_id",
        "Description": "Normalized business concept assigned to the column, used to group similar fields across tables.",
    },
    {
        "Field": "concept_confidence",
        "Description": "Confidence score for the selected concept_id, typically between 0 and 1.",
    },
    {
        "Field": "concept_alias_group",
        "Description": "Normalized alias bucket used to compare similar column names across tables.",
    },
    {
        "Field": "concept_evidence_json",
        "Description": "Serialized JSON showing the signals that supported the chosen concept.",
    },
    {
        "Field": "concept_sources_json",
        "Description": "Serialized JSON list of source types used during concept inference, such as name, values, or profile.",
    },
    {
        "Field": "unit",
        "Description": "Detected source unit for the column, usually taken from the name or stored unit metadata.",
    },
    {
        "Field": "unit_source",
        "Description": "Where the unit came from, for example name-based detection or manually provided metadata.",
    },
    {
        "Field": "canonical_unit",
        "Description": "Normalized target unit used for consistent comparisons and conversions.",
    },
    {
        "Field": "unit_system",
        "Description": "Measurement system associated with the unit, such as metric or imperial.",
    },
    {
        "Field": "unit_confidence",
        "Description": "Confidence level for the detected unit assignment.",
    },
    {
        "Field": "unit_notes",
        "Description": "Additional notes about unit detection or normalization behavior.",
    },
    {
        "Field": "factor_to_canonical",
        "Description": "Multiplication factor used to convert the source unit into the canonical unit.",
    },
    {
        "Field": "offset_to_canonical",
        "Description": "Additive offset used during conversion to the canonical unit.",
    },
    {
        "Field": "conversion_formula",
        "Description": "Human-readable formula showing how to convert the source unit to the canonical unit.",
    },
]

_CLASSIFICATION_COLUMN_PREFIX = "om_class_"


def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dict):
        pairs = [f"{k}={_cell_value(v)}" for k, v in value.items()]
        return "; ".join(pairs)
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(_cell_value(v)) for v in value)
    return value


def _is_nested(value):
    return isinstance(value, (dict, list))


def _flatten_dict(value, prefix=""):
    out = {}
    if not isinstance(value, dict):
        return out
    for key, item in value.items():
        col = f"{prefix}{key}" if not prefix else f"{prefix}_{key}"
        if isinstance(item, dict):
            out.update(_flatten_dict(item, col))
        elif isinstance(item, list):
            out[col] = ", ".join(str(_cell_value(v)) for v in item)
        else:
            out[col] = _cell_value(item)
    return out


def _split_nested_sections(sheet_name, rows):
    if not rows:
        return [("Data", rows)]

    nested_fields = []
    for key in rows[0].keys():
        if any(_is_nested(row.get(key)) for row in rows):
            nested_fields.append(key)

    if not nested_fields:
        return [("Data", rows)]

    main_rows = []
    for row in rows:
        new_row = deepcopy(row)
        for field in nested_fields:
            nested_value = row.get(field)
            if isinstance(nested_value, list):
                new_row[f"{field}_count"] = len(nested_value)
            elif isinstance(nested_value, dict):
                new_row[f"{field}_count"] = len(nested_value.keys())
            else:
                new_row[f"{field}_count"] = 0
            new_row.pop(field, None)
        main_rows.append(new_row)

    scalar_fields = [k for k in rows[0].keys() if k not in nested_fields]
    child_sections = []

    for field in nested_fields:
        child_rows = []
        for row_idx, row in enumerate(rows, start=1):
            nested_value = row.get(field)
            if nested_value is None or nested_value == {} or nested_value == []:
                continue

            base = {"parent_row": row_idx}
            for key in scalar_fields:
                value = row.get(key)
                if not _is_nested(value):
                    base[key] = _cell_value(value)

            if isinstance(nested_value, dict):
                child = dict(base)
                child.update(_flatten_dict(nested_value))
                child_rows.append(child)
            elif isinstance(nested_value, list):
                for item_idx, item in enumerate(nested_value, start=1):
                    child = dict(base)
                    child["item_index"] = item_idx
                    if isinstance(item, dict):
                        child.update(_flatten_dict(item))
                    elif isinstance(item, list):
                        child["value"] = ", ".join(str(_cell_value(v)) for v in item)
                    else:
                        child["value"] = _cell_value(item)
                    child_rows.append(child)
            else:
                child = dict(base)
                child["value"] = _cell_value(nested_value)
                child_rows.append(child)

        if child_rows:
            child_sections.append((f"{sheet_name}.{field}", child_rows))

    return [("Data", main_rows)] + child_sections


def _join_list(values):
    if not values:
        return ""
    return ", ".join(str(v) for v in values)


def _safe_classification_key(value):
    text = str(value or "").strip()
    text = "".join(ch if ch.isalnum() else "_" for ch in text)
    text = text.strip("_")
    return text or "Classification"


def _classification_excel_column_name(classification_name):
    return f"{_CLASSIFICATION_COLUMN_PREFIX}{_safe_classification_key(classification_name)}"


def _excel_classifications(metadata, entity_level):
    classifications = []
    for classification in metadata.get("openmetadata_classifications", []) or []:
        if not isinstance(classification, dict):
            continue
        name = str(classification.get("name", "")).strip()
        if not name:
            continue
        if entity_level == "table" and name.upper() == "PII":
            continue
        classifications.append(classification)
    return classifications


def _classification_selection_fields(tags, metadata, entity_level):
    selected = {}
    selected_tags = set(tags or [])
    for classification in _excel_classifications(metadata, entity_level):
        header = _classification_excel_column_name(classification.get("name", ""))
        selected_value = ""
        for option in classification.get("options", []) or []:
            if not isinstance(option, dict):
                continue
            fqn = str(option.get("fqn", "")).strip()
            if fqn and fqn in selected_tags:
                selected_value = fqn
                break
        selected[header] = selected_value
    return selected


def _flatten_sensitive_fields(values):
    if not isinstance(values, dict) or not values:
        return ""
    items = [f"{k}:{v}" for k, v in values.items()]
    return "; ".join(items)


def _unit_fields(column):
    unit_context = column.get("unit_context") or {}
    conversion = unit_context.get("conversion") or {}

    detected_unit = column.get("unit") or unit_context.get("detected_unit", "")
    unit_source = column.get("unit_source") or unit_context.get("detection_source", "")

    return {
        "unit": detected_unit or "",
        "unit_source": unit_source or "",
        "canonical_unit": unit_context.get("canonical_unit", ""),
        "unit_system": unit_context.get("unit_system", ""),
        "unit_confidence": unit_context.get("detection_confidence", ""),
        "unit_notes": unit_context.get("notes", ""),
        "factor_to_canonical": conversion.get("factor_to_canonical", ""),
        "offset_to_canonical": conversion.get("offset_to_canonical", ""),
        "conversion_formula": conversion.get("formula", ""),
    }


def _compact_json(value):
    if value in (None, "", [], {}):
        return ""
    try:
        return json.dumps(value, ensure_ascii=True, separators=(",", ":"))
    except Exception:
        return _cell_value(value)


def _contacts_rows(source_system_context):
    contacts = source_system_context.get("contacts", [])
    rows = []
    if isinstance(contacts, list):
        for c in contacts:
            if isinstance(c, dict):
                rows.append(
                    {
                        "contact_name": c.get("name", ""),
                        "role": c.get("role", ""),
                        "email": c.get("email", ""),
                        "phone": c.get("phone", ""),
                        "notes": c.get("notes", ""),
                    }
                )
            else:
                rows.append(
                    {
                        "contact_name": _cell_value(c),
                        "role": "",
                        "email": "",
                        "phone": "",
                        "notes": "",
                    }
                )
    # Provide blank fillable lines for manual completion.
    while len(rows) < 8:
        rows.append({"contact_name": "", "role": "", "email": "", "phone": "", "notes": ""})
    return rows


def _delete_management_rows(source_system_context):
    instruction = source_system_context.get("delete_management_instruction", "")
    rows = [
        {
            "table_name": "",
            "delete_strategy": "",
            "instruction": _cell_value(instruction),
            "notes": "",
        }
    ]
    while len(rows) < 8:
        rows.append({"table_name": "", "delete_strategy": "", "instruction": "", "notes": ""})
    return rows


def _restrictions_rows(source_system_context):
    restrictions = source_system_context.get("restrictions", "")
    rows = []
    if isinstance(restrictions, list):
        for r in restrictions:
            if isinstance(r, dict):
                rows.append(
                    {
                        "table_name": r.get("table_name", ""),
                        "restriction_type": r.get("type", ""),
                        "scope": r.get("scope", ""),
                        "details": r.get("details", ""),
                        "owner": r.get("owner", ""),
                    }
                )
            else:
                rows.append(
                    {
                        "table_name": "",
                        "restriction_type": "",
                        "scope": "",
                        "details": _cell_value(r),
                        "owner": "",
                    }
                )
    elif isinstance(restrictions, dict):
        rows.append(
            {
                "table_name": restrictions.get("table_name", ""),
                "restriction_type": restrictions.get("type", ""),
                "scope": restrictions.get("scope", ""),
                "details": restrictions.get("details", ""),
                "owner": restrictions.get("owner", ""),
            }
        )
    else:
        rows.append({"table_name": "", "restriction_type": "", "scope": "", "details": _cell_value(restrictions), "owner": ""})

    while len(rows) < 8:
        rows.append({"table_name": "", "restriction_type": "", "scope": "", "details": "", "owner": ""})
    return rows


def _late_arriving_manual_rows(source_system_context):
    rows = [
        {
            "table_name": "",
            "business_date_column": "",
            "system_ts_column": "",
            "lookback_days": "",
            "policy_notes": _cell_value(source_system_context.get("late_arriving_data_manual", "")),
        }
    ]
    while len(rows) < 8:
        rows.append(
            {
                "table_name": "",
                "business_date_column": "",
                "system_ts_column": "",
                "lookback_days": "",
                "policy_notes": "",
            }
        )
    return rows


def _volume_projection_manual_rows(source_system_context):
    rows = [
        {
            "entity_scope": "",
            "projection_horizon_months": "",
            "growth_assumption_pct": "",
            "basis": "",
            "notes": _cell_value(source_system_context.get("volume_size_projection_manual", "")),
        }
    ]
    while len(rows) < 8:
        rows.append(
            {
                "entity_scope": "",
                "projection_horizon_months": "",
                "growth_assumption_pct": "",
                "basis": "",
                "notes": "",
            }
        )
    return rows


def _db_analysis_config_rows(source_system_context):
    cfg = source_system_context.get("db_analysis_config", {}) or {}
    rows = [
        {
            "exclude_schemas": _join_list(cfg.get("exclude_schemas", [])),
            "exclude_tables": _join_list(cfg.get("exclude_tables", [])),
            "max_row_limit": _cell_value(cfg.get("max_row_limit", "")),
        }
    ]
    while len(rows) < 3:
        rows.append(
            {
                "exclude_schemas": "",
                "exclude_tables": "",
                "max_row_limit": "",
            }
        )
    return rows


def _openmetadata_classification_rows(metadata):
    rows = []
    for classification in metadata.get("openmetadata_classifications", []) or []:
        if not isinstance(classification, dict):
            continue
        options = classification.get("options") or []
        if not options:
            rows.append(
                {
                    "classification_name": classification.get("name", ""),
                    "provider": classification.get("provider", ""),
                    "mutually_exclusive": classification.get("mutually_exclusive", ""),
                    "allowed_on": _join_list(classification.get("allowed_on", [])),
                    "option_name": "",
                    "option_fqn": "",
                }
            )
            continue
        for option in options:
            if not isinstance(option, dict):
                continue
            rows.append(
                {
                    "classification_name": classification.get("name", ""),
                    "provider": classification.get("provider", ""),
                    "mutually_exclusive": classification.get("mutually_exclusive", ""),
                    "allowed_on": _join_list(classification.get("allowed_on", [])),
                    "option_name": option.get("name", ""),
                    "option_fqn": option.get("fqn", ""),
                }
            )
    return rows


def _classification_validation_rows(metadata):
    classifications = _excel_classifications(metadata, "column")
    if not classifications:
        return []
    headers = [_classification_excel_column_name(classification.get("name", "")) for classification in classifications]
    option_lists = []
    max_options = 0
    for classification in classifications:
        options = []
        for option in classification.get("options", []) or []:
            if not isinstance(option, dict):
                continue
            fqn = str(option.get("fqn", "")).strip()
            if fqn:
                options.append(fqn)
        option_lists.append(options)
        max_options = max(max_options, len(options))
    rows = []
    for idx in range(max_options):
        row = {}
        for header, options in zip(headers, option_lists):
            row[header] = options[idx] if idx < len(options) else ""
        rows.append(row)
    return rows


def _business_terms_sheet_rows(glossary_payload):
    if not glossary_payload:
        return []
    glossaries = glossary_payload.get("glossaries") or []
    glossary_display = {}
    for g in glossaries:
        if isinstance(g, dict):
            fqn = g.get("fullyQualifiedName") or g.get("name", "")
            glossary_display[fqn] = g.get("displayName") or g.get("name", "")
    rows = []
    for term in glossary_payload.get("terms") or []:
        if not isinstance(term, dict):
            continue
        glossary_info = term.get("glossary") or {}
        glossary_fqn = glossary_info.get("fullyQualifiedName") or glossary_info.get("name", "")
        glossary_name = glossary_display.get(glossary_fqn, "") or glossary_info.get("displayName") or glossary_info.get("name", "")
        rows.append(
            {
                "glossary": glossary_name,
                "term": term.get("displayName") or term.get("name", ""),
                "fully_qualified_name": term.get("fullyQualifiedName", ""),
                "description": term.get("description", ""),
                "synonyms": _join_list(term.get("synonyms", [])),
                "status": term.get("entityStatus", ""),
            }
        )
    rows.sort(key=lambda r: (r["glossary"], r["term"]))
    return rows


def _glossary_sheet_rows(metadata):
    rows = list(_GLOSSARY_ROWS)
    for classification in metadata.get("openmetadata_classifications", []) or []:
        if not isinstance(classification, dict):
            continue
        classification_name = str(classification.get("name", "")).strip()
        if not classification_name:
            continue
        rows.append(
            {
                "Field": f"classification: {classification_name}",
                "Description": str(classification.get("description") or "").strip(),
            }
        )
        for option in classification.get("options", []) or []:
            if not isinstance(option, dict):
                continue
            option_fqn = str(option.get("fqn", "")).strip()
            if not option_fqn:
                continue
            rows.append(
                {
                    "Field": f"classification option: {option_fqn}",
                    "Description": str(option.get("description") or "").strip(),
                }
            )
    return rows


def _row_from_finding(schema_name, table_name, idx, finding):
    if not isinstance(finding, dict):
        return {
            "schema": schema_name,
            "table_name": table_name,
            "finding_index": idx,
            "check": "",
            "severity": "",
            "column": "",
            "detail": _cell_value(finding),
            "recommendation": "",
            "distinct_values": "",
            "suggested_domain": "",
            "sample_values": "",
            "cardinality": "",
            "delete_strategy": "",
            "soft_delete_column": "",
            "soft_delete_type": "",
            "has_audit_trail": "",
            "business_date_column": "",
            "system_ts_column": "",
            "server_timezone": "",
            "timezone_columns": "",
            "distinct_timezones": "",
            "tz_aware_count": "",
            "tz_naive_count": "",
            "detected_unit": "",
            "canonical_unit": "",
            "extra_json": "",
        }

    lag_stats = finding.get("lag_stats") or {}
    if not isinstance(lag_stats, dict):
        lag_stats = {}

    known_keys = {
        "check", "severity", "column", "detail", "recommendation",
        "distinct_values", "suggested_domain", "sample_values", "cardinality",
        "delete_strategy", "soft_delete_column", "soft_delete_type", "has_audit_trail",
        "business_date_column", "system_ts_column", "recommended_lookback_days", "lag_stats",
        "server_timezone", "columns", "distinct_timezones", "tz_aware_count", "tz_naive_count",
        "detected_unit", "canonical_unit",
    }
    extra_fields = {k: v for k, v in finding.items() if k not in known_keys}

    return {
        "schema": schema_name,
        "table_name": table_name,
        "finding_index": idx,
        "check": finding.get("check", ""),
        "severity": finding.get("severity", ""),
        "column": finding.get("column", ""),
        "detail": finding.get("detail", ""),
        "recommendation": finding.get("recommendation", ""),
        "distinct_values": _cell_value(finding.get("distinct_values", "")),
        "suggested_domain": _cell_value(finding.get("suggested_domain", "")),
        "sample_values": _cell_value(finding.get("sample_values", "")),
        "cardinality": finding.get("cardinality", ""),
        "delete_strategy": finding.get("delete_strategy", ""),
        "soft_delete_column": finding.get("soft_delete_column", ""),
        "soft_delete_type": finding.get("soft_delete_type", ""),
        "has_audit_trail": _cell_value(finding.get("has_audit_trail", "")),
        "business_date_column": finding.get("business_date_column", ""),
        "system_ts_column": finding.get("system_ts_column", ""),
        "server_timezone": finding.get("server_timezone", ""),
        "timezone_columns": _compact_json(finding.get("columns")),
        "distinct_timezones": _cell_value(finding.get("distinct_timezones", "")),
        "tz_aware_count": finding.get("tz_aware_count", ""),
        "tz_naive_count": finding.get("tz_naive_count", ""),
        "detected_unit": finding.get("detected_unit", ""),
        "canonical_unit": finding.get("canonical_unit", ""),
        "extra_json": _compact_json(extra_fields),
    }


def _derive_database(connection, metadata):
    db = connection.get("database")
    if db:
        return db
    db_url = metadata.get("database_url") or ""
    if not db_url:
        return ""
    parsed = urlparse(db_url if "://" in db_url else f"dummy://{db_url}")
    if parsed.path and parsed.path != "/":
        return parsed.path.lstrip("/")
    q = parse_qs(parsed.query or "")
    service = q.get("service_name", [""])[0]
    if service:
        return service
    return ""


def _sheet_name(base_name, used_names):
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(base_name or "Table"))
    cleaned = cleaned.strip("'").strip() or "Table"
    candidate = cleaned[:31] or "Table"
    counter = 1
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[: max(0, 31 - len(suffix))]}{suffix}" or f"Table{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def _table_overview_row(table, metadata):
    row = {
        "schema": table.get("schema", ""),
        "table_name": table.get("table", ""),
        "row_count": table.get("row_count", ""),
        "has_primary_key": table.get("has_primary_key", ""),
        "has_foreign_keys": table.get("has_foreign_keys", ""),
        "has_sensitive_fields": table.get("has_sensitive_fields", ""),
        "cdc_enabled": table.get("cdc_enabled", ""),
        "table_description": table.get("table_description", ""),
        "glossary_terms": _join_list(table.get("glossary_terms", [])),
        "classification_summary_json": _compact_json(table.get("classification_summary")),
        "unit_summary_json": _compact_json(table.get("unit_summary")),
        "row_count_projection_1y": table.get("row_count_projection_1y", ""),
        "row_count_projection_2y": table.get("row_count_projection_2y", ""),
        "row_count_projection_5y": table.get("row_count_projection_5y", ""),
    }
    row.update(_classification_selection_fields(table.get("classification_tags", []), metadata, "table"))
    return row


def _simple_value_rows(values, key_name):
    rows = []
    for idx, value in enumerate(values or [], start=1):
        rows.append(
            {
                "item_index": idx,
                key_name: _cell_value(value),
            }
        )
    return rows


def _mapping_rows(mapping, value_name):
    if not isinstance(mapping, dict):
        return []
    return [
        {"column_name": key, value_name: _cell_value(value)}
        for key, value in mapping.items()
    ]


def _column_rows(table, metadata):
    rows = []
    field_classifications = table.get("field_classifications") or {}
    sensitive_fields = table.get("sensitive_fields") or {}
    for column in table.get("columns", []) or []:
        unit_data = _unit_fields(column)
        column_name = column.get("name", "")
        row = {
            "column_name": column_name,
            "data_type": column.get("type", ""),
            "nullable": column.get("nullable", ""),
            "classification": _cell_value(field_classifications.get(column_name, "")),
            "sensitivity_label": _cell_value(sensitive_fields.get(column_name, "")),
            "cardinality": column.get("cardinality", ""),
            "null_count": column.get("null_count", ""),
            "data_category": column.get("data_category", ""),
            "semantic_class": column.get("semantic_class", ""),
            "column_description": column.get("column_description", column.get("description", "")),
            "glossary_terms": _join_list(column.get("glossary_terms", [])),
            "concept_id": column.get("concept_id", ""),
            "concept_confidence": column.get("concept_confidence", ""),
            "concept_alias_group": column.get("concept_alias_group", ""),
            "concept_evidence_json": _compact_json(column.get("concept_evidence")),
            "concept_sources_json": _compact_json(column.get("concept_sources")),
            "unit": unit_data["unit"],
            "unit_source": unit_data["unit_source"],
            "canonical_unit": unit_data["canonical_unit"],
            "unit_system": unit_data["unit_system"],
            "unit_confidence": unit_data["unit_confidence"],
            "unit_notes": unit_data["unit_notes"],
            "factor_to_canonical": unit_data["factor_to_canonical"],
            "offset_to_canonical": unit_data["offset_to_canonical"],
            "conversion_formula": unit_data["conversion_formula"],
            "range_min": (column.get("data_range") or {}).get("min", ""),
            "range_max": (column.get("data_range") or {}).get("max", ""),
        }
        row.update(_classification_selection_fields(column.get("classification_tags", []), metadata, "column"))
        rows.append(row)
    return rows


def _join_candidate_rows(table):
    rows = []
    for candidate in table.get("join_candidates", []) or []:
        if isinstance(candidate, dict):
            rows.append(
                {
                    "column_name": candidate.get("column", ""),
                    "target_table": candidate.get("target_table", ""),
                    "target_column": candidate.get("target_column", ""),
                    "confidence": candidate.get("confidence", ""),
                }
            )
        else:
            rows.append(
                {
                    "column_name": "",
                    "target_table": _cell_value(candidate),
                    "target_column": "",
                    "confidence": "",
                }
            )
    return rows


def _foreign_key_rows(table):
    return [
        {
            "column_name": fk.get("column", ""),
            "references": fk.get("references", ""),
        }
        for fk in (table.get("foreign_keys", []) or [])
        if isinstance(fk, dict)
    ]


def _sample_data_rows(table):
    rows = []
    for sample_col, values in (table.get("sample_data", {}) or {}).items():
        if not isinstance(values, list):
            values = [values]
        for idx, value in enumerate(values, start=1):
            rows.append(
                {
                    "sample_column": sample_col,
                    "sample_index": idx,
                    "sample_value": _cell_value(value),
                }
            )
    return rows


def _table_findings_rows(table):
    schema_name = table.get("schema", "")
    table_name = table.get("table", "")
    return [
        _row_from_finding(schema_name, table_name, idx, finding)
        for idx, finding in enumerate(((table.get("data_quality") or {}).get("findings") or []), start=1)
    ]


def _source_system_sections(metadata, connection, source_system_context):
    metadata_rows = [
        {"property": key, "value": _cell_value(value)}
        for key, value in metadata.items()
        if key != "openmetadata_classifications"
    ]
    connection_rows = [{"property": key, "value": _cell_value(value)} for key, value in connection.items()]
    sections = [
        ("Metadata", metadata_rows or [{"property": "", "value": ""}]),
        ("Connection", connection_rows or [{"property": "", "value": ""}]),
    ]
    openmetadata_rows = _openmetadata_classification_rows(metadata)
    if openmetadata_rows:
        sections.append(("OpenMetadataClassifications", openmetadata_rows))
    sections.extend(
        [
            ("ContactsManual", _contacts_rows(source_system_context)),
            ("DeleteManagementManual", _delete_management_rows(source_system_context)),
            ("RestrictionsManual", _restrictions_rows(source_system_context)),
            ("LateArrivingDataManual", _late_arriving_manual_rows(source_system_context)),
            ("VolumeSizeProjectionManual", _volume_projection_manual_rows(source_system_context)),
            ("DbAnalysisConfig", _db_analysis_config_rows(source_system_context)),
        ]
    )
    return sections


def _table_sheet_sections(table, metadata):
    sections = [("Overview", [_table_overview_row(table, metadata)])]

    primary_keys = _simple_value_rows(table.get("primary_keys", []), "column_name")
    if primary_keys:
        sections.append(("PrimaryKeys", primary_keys))

    foreign_keys = _foreign_key_rows(table)
    if foreign_keys:
        sections.append(("ForeignKeys", foreign_keys))

    incremental = _simple_value_rows(table.get("incremental_columns", []), "column_name")
    if incremental:
        sections.append(("IncrementalColumns", incremental))

    partition = _simple_value_rows(table.get("partition_columns", []), "column_name")
    if partition:
        sections.append(("PartitionColumns", partition))

    partition_candidates = _simple_value_rows(table.get("partition_columns_candidates", []), "column_name")
    if partition_candidates:
        sections.append(("PartitionColumnCandidates", partition_candidates))

    join_candidates = _join_candidate_rows(table)
    if join_candidates:
        sections.append(("JoinCandidates", join_candidates))

    columns = _column_rows(table, metadata)
    if columns:
        sections.append(("Columns", columns))

    sample_data = _sample_data_rows(table)
    if sample_data:
        sections.append(("SampleData", sample_data))

    findings = _table_findings_rows(table)
    if findings:
        sections.append(("DataQualityFindings", findings))

    return sections


def _collect_sheets(payload, glossary_payload=None):
    tables = payload.get("tables", []) or []
    connection = payload.get("connection", {}) or {}
    metadata = payload.get("metadata", {}) or {}
    data_quality_summary = payload.get("data_quality_summary", {}) or {}
    source_system_context = payload.get("source_system_context", {}) or {}
    first_table_schema = tables[0].get("schema", "") if tables else ""

    source_name = (
        connection.get("source_name")
        or connection.get("host")
        or metadata.get("database_url")
        or ""
    )
    source_type = (
        connection.get("source_type")
        or connection.get("driver")
        or ""
    )
    schema_name = (
        connection.get("schema")
        or metadata.get("schema_filter")
        or first_table_schema
        or ""
    )

    summary = [
        {"metric": "source_name", "value": source_name},
        {"metric": "source_type", "value": source_type},
        {"metric": "database", "value": _derive_database(connection, metadata)},
        {"metric": "schema", "value": schema_name},
        {"metric": "port", "value": connection.get("port", "")},
        {"metric": "database_timezone", "value": connection.get("timezone", "")},
        {"metric": "tables_count", "value": len(tables)},
    ]
    dq_by_check_rows = []
    dq_constraints_rows = []
    dq_other_rows = []

    for key, value in data_quality_summary.items():
        if isinstance(value, dict):
            if key == "by_check":
                for check_name, check_value in value.items():
                    dq_by_check_rows.append({"check": check_name, "count": _cell_value(check_value)})
            elif key == "constraints_found":
                for constraint_name, constraint_value in value.items():
                    dq_constraints_rows.append(
                        {"constraint": constraint_name, "count": _cell_value(constraint_value)}
                    )
            else:
                for item_name, item_value in value.items():
                    dq_other_rows.append(
                        {
                            "metric_group": f"data_quality_{key}",
                            "item": item_name,
                            "value": _cell_value(item_value),
                        }
                    )
        else:
            summary.append({"metric": f"data_quality_{key}", "value": _cell_value(value)})

    data_quality_findings_rows = []

    for table in tables:
        schema_name = table.get("schema", "")
        table_name = table.get("table", "")
        findings = ((table.get("data_quality") or {}).get("findings") or [])
        for idx, finding in enumerate(findings, start=1):
            data_quality_findings_rows.append(
                _row_from_finding(schema_name, table_name, idx, finding)
            )

    summary_sections = [("Overview", summary)]
    if dq_by_check_rows:
        summary_sections.append(("DataQualityByCheck", dq_by_check_rows))
    if dq_constraints_rows:
        summary_sections.append(("DataQualityConstraints", dq_constraints_rows))
    if dq_other_rows:
        summary_sections.append(("DataQualityDetails", dq_other_rows))

    sheets = {
        "Summary": {"sections": summary_sections},
        "SourceSystem": {"sections": _source_system_sections(metadata, connection, source_system_context)},
        "DataQualityFindings": data_quality_findings_rows,
        "Glossary": _glossary_sheet_rows(metadata),
    }
    business_terms = _business_terms_sheet_rows(glossary_payload)
    if business_terms:
        sheets["DataGovernanceTerms"] = business_terms
    classification_validation_rows = _classification_validation_rows(metadata)
    if classification_validation_rows:
        sheets["__dv_classifications"] = classification_validation_rows
    used_sheet_names = set(sheets.keys())
    for table in tables:
        sheets[_sheet_name(table.get("table", "Table"), used_sheet_names)] = {
            "sections": _table_sheet_sections(table, metadata)
        }
    return sheets


def _apply_classification_validations(wb):
    if "__dv_classifications" not in wb.sheetnames:
        return
    options_ws = wb["__dv_classifications"]
    option_headers = [cell.value for cell in options_ws[1] if cell.value not in (None, "")]
    if not option_headers:
        return
    option_ranges = {}
    for idx, header in enumerate(option_headers, start=1):
        col_letter = get_column_letter(idx)
        option_ranges[str(header)] = f"'__dv_classifications'!${col_letter}$2:${col_letter}${options_ws.max_row}"

    for ws in wb.worksheets:
        if ws.title in {"Summary", "SourceSystem", "DataQualityFindings", "Glossary", "DataGovernanceTerms", "__dv_classifications"}:
            continue
        if ws.sheet_state != "visible":
            continue
        rows = list(ws.iter_rows())
        i = 0
        while i < len(rows):
            row = rows[i] or ()
            section_name = row[0].value if row and len(row) > 0 else None
            if section_name in (None, ""):
                i += 1
                continue
            if i + 1 >= len(rows):
                break
            headers = [cell.value for cell in rows[i + 1]]
            data_start = i + 2
            data_end = data_start - 1
            j = data_start
            while j < len(rows):
                current = rows[j]
                if all(cell.value in (None, "") for cell in current):
                    break
                data_end = j
                j += 1
            if data_end >= data_start:
                for col_idx, header in enumerate(headers, start=1):
                    if header not in option_ranges:
                        continue
                    dv = DataValidation(
                        type="list",
                        formula1=option_ranges[str(header)],
                        allow_blank=True,
                        showErrorMessage=True,
                        errorStyle="stop",
                    )
                    wb[ws.title].add_data_validation(dv)
                    for row_idx in range(data_start, data_end + 1):
                        dv.add(ws.cell(row=row_idx + 1, column=col_idx))
            i = j + 1


def _chunk_text(value, size=30000):
    if not value:
        return []
    return [value[i : i + size] for i in range(0, len(value), size)]


def _roundtrip_sheets(payload):
    # Store the complete original payload in hidden sheets so reverse conversion
    # can preserve fields that are not represented in user-facing tabs.
    raw = json.dumps(payload, ensure_ascii=True, separators=(",", ":"))
    chunks = _chunk_text(raw)
    return {
        "__rt_meta": [
            {"key": "format_version", "value": "2"},
            {"key": "payload_chunks", "value": len(chunks)},
        ],
        "__rt_payload": [
            {"chunk_index": i + 1, "payload_chunk": chunk}
            for i, chunk in enumerate(chunks)
        ],
    }


def _style_sheet(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)


def _write_sheet(ws, rows):
    headers = list(rows[0].keys()) if rows else ["note"]
    ws.append(headers)
    if rows:
        for row in rows:
            ws.append([_cell_value(row.get(h, "")) for h in headers])
    else:
        ws.append(["No rows"])
    _style_sheet(ws)
    if ws.title == "DataQualityFindings":
        wrap_cols = {"detail", "recommendation", "timezone_columns", "extra_json"}
        for idx, header in enumerate(headers, start=1):
            if header in wrap_cols:
                col_letter = ws.cell(row=1, column=idx).column_letter
                ws.column_dimensions[col_letter].width = 60
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=idx)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _write_multi_section_sheet(ws, sections):
    row_idx = 1
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(color="1F4E78", bold=True)

    for section_name, rows in sections:
        ws.cell(row=row_idx, column=1, value=section_name)
        ws.cell(row=row_idx, column=1).fill = section_fill
        ws.cell(row=row_idx, column=1).font = section_font
        row_idx += 1

        headers = list(rows[0].keys()) if rows else ["note"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=header)
        for cell in ws[row_idx]:
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1

        if rows:
            for row in rows:
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=_cell_value(row.get(header, "")))
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value="No rows")
            row_idx += 1

        row_idx += 1

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)
    if ws.title == "SourceSystem":
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Wider manual input fields for easier data entry.
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 64
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 36


def _write_workbook(sheet_rows, output_path):
    wb = Workbook()
    first = True
    for sheet_name, rows in sheet_rows.items():
        if isinstance(rows, dict) and isinstance(rows.get("sections"), list):
            has_data = any(section_rows for _, section_rows in rows["sections"])
            if not has_data:
                continue
        elif isinstance(rows, list) and not rows:
            continue

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        if isinstance(rows, dict) and isinstance(rows.get("sections"), list):
            _write_multi_section_sheet(ws, rows["sections"])
        else:
            sections = _split_nested_sections(sheet_name, rows)
            if len(sections) == 1 and sections[0][0] == "Data":
                _write_sheet(ws, sections[0][1])
            else:
                _write_multi_section_sheet(ws, sections)
        if sheet_name.startswith("__rt_") or sheet_name.startswith("__dv_"):
            ws.sheet_state = "hidden"
    if first:
        ws = wb.active
        ws.title = "Summary"
        _write_sheet(ws, [])
    _apply_classification_validations(wb)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert schema JSON to styled Excel workbook")
    parser.add_argument("input_json", help="Path to source JSON file")
    parser.add_argument("output_xlsx", nargs="?", help="Path to output .xlsx file")
    parser.add_argument(
        "--glossary-json",
        dest="glossary_json",
        default=None,
        help="Path to OpenMetadata glossary JSON file. When omitted the script fetches directly from OpenMetadata.",
    )
    parser.add_argument(
        "--no-openmetadata",
        dest="no_openmetadata",
        action="store_true",
        default=False,
        help="Skip automatic OpenMetadata glossary fetch (omit the DataGovernanceTerms sheet).",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    input_path = Path(args.input_json).expanduser().resolve()
    output_path = (
        Path(args.output_xlsx).expanduser().resolve()
        if args.output_xlsx
        else input_path.with_suffix(".xlsx")
    )

    payload = json.loads(input_path.read_text(encoding="utf-8"))

    glossary_payload = None
    if args.glossary_json:
        glossary_path = Path(args.glossary_json).expanduser().resolve()
        glossary_payload = json.loads(glossary_path.read_text(encoding="utf-8"))
    elif not args.no_openmetadata:
        glossary_payload = _om_fetch_glossary_payload()

    sheet_rows = _collect_sheets(payload, glossary_payload=glossary_payload)
    sheet_rows.update(_roundtrip_sheets(payload))
    _write_workbook(sheet_rows, output_path)
    row_count = sum(len(v) for v in sheet_rows.values())
    print(f"Wrote {row_count} rows across {len(sheet_rows)} sheets to {output_path}")


if __name__ == "__main__":
    main()
