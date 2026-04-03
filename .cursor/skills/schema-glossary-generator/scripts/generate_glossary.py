#!/usr/bin/env python3
import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_VALUES = ["draft", "approved", "rejected"]

KEYWORD_LIBRARY = [
    {
        "term": "Customer",
        "term_type": "business_entity",
        "keywords": ["customer", "consumer", "buyer", "client"],
        "definition": "A person or organization that purchases goods or services from the business.",
        "business_usage": "Used across sales, service, fulfillment, and reporting processes.",
        "synonyms": ["Client", "Buyer"],
    },
    {
        "term": "Supplier",
        "term_type": "business_entity",
        "keywords": ["supplier", "vendor"],
        "definition": "A party that provides goods or services to the business.",
        "business_usage": "Used in procurement, replenishment, and payables processes.",
        "synonyms": ["Vendor"],
    },
    {
        "term": "Product",
        "term_type": "business_entity",
        "keywords": ["product", "item", "sku", "goods", "merchandise"],
        "definition": "A sellable or purchasable good managed by the business.",
        "business_usage": "Used in catalog, pricing, inventory, order, and reporting workflows.",
        "synonyms": ["Item", "SKU"],
    },
    {
        "term": "Inventory",
        "term_type": "business_entity",
        "keywords": ["inventory", "stock", "warehouse"],
        "definition": "The quantity of products available for sale, use, or replenishment.",
        "business_usage": "Used to track stock availability, replenishment needs, and fulfillment readiness.",
        "synonyms": ["Stock"],
    },
    {
        "term": "Sales Order",
        "term_type": "business_process",
        "keywords": ["sales order", "order", "order-to-cash", "sale"],
        "definition": "A commercial transaction that records a customer's requested purchase.",
        "business_usage": "Used to manage order capture, pricing, fulfillment, billing, and revenue reporting.",
        "synonyms": ["Order"],
    },
    {
        "term": "Purchase Order",
        "term_type": "business_process",
        "keywords": ["purchase order", "procurement", "purchasing", "replenishment"],
        "definition": "A transaction that records goods or services ordered from a supplier.",
        "business_usage": "Used to manage supplier commitments, inbound deliveries, and spend tracking.",
        "synonyms": ["PO"],
    },
    {
        "term": "Invoice",
        "term_type": "business_process",
        "keywords": ["invoice", "billing", "bill"],
        "definition": "A financial document requesting payment for delivered goods or services.",
        "business_usage": "Used in billing, receivables, collections, and revenue reconciliation.",
        "synonyms": ["Bill"],
    },
    {
        "term": "Payment",
        "term_type": "business_process",
        "keywords": ["payment", "pay", "collection", "settlement"],
        "definition": "The transfer of funds to settle a financial obligation.",
        "business_usage": "Used in cash application, collections, refunds, and treasury reporting.",
        "synonyms": ["Settlement"],
    },
    {
        "term": "Delivery",
        "term_type": "business_process",
        "keywords": ["delivery", "shipping", "shipment", "dispatch", "fulfillment"],
        "definition": "The movement of ordered goods to the customer or destination.",
        "business_usage": "Used in fulfillment planning, logistics execution, and service-level reporting.",
        "synonyms": ["Shipment", "Fulfillment"],
    },
    {
        "term": "Return",
        "term_type": "business_process",
        "keywords": ["return", "refund", "reverse logistics"],
        "definition": "A process that handles goods sent back after a sale or delivery.",
        "business_usage": "Used in customer service, inventory adjustment, and financial reconciliation.",
        "synonyms": ["Refund"],
    },
    {
        "term": "Order Status",
        "term_type": "status",
        "keywords": ["status", "order status", "lifecycle"],
        "definition": "The current stage of an order or transaction in its lifecycle.",
        "business_usage": "Used to monitor process progression, exceptions, and operational reporting.",
        "synonyms": ["Lifecycle Status"],
    },
    {
        "term": "Unit Price",
        "term_type": "business_measure",
        "keywords": ["price", "unit price", "pricing"],
        "definition": "The amount charged for a single unit of a product or service.",
        "business_usage": "Used in quoting, ordering, billing, and margin analysis.",
        "synonyms": ["Price"],
    },
    {
        "term": "Quantity",
        "term_type": "business_measure",
        "keywords": ["quantity", "volume", "units"],
        "definition": "The count or amount of a product or service involved in a transaction.",
        "business_usage": "Used in ordering, inventory, fulfillment, and productivity reporting.",
        "synonyms": ["Units"],
    },
    {
        "term": "Total Amount",
        "term_type": "business_measure",
        "keywords": ["amount", "total", "revenue", "sales"],
        "definition": "The full monetary value associated with a transaction or business event.",
        "business_usage": "Used in financial reporting, billing, and performance analysis.",
        "synonyms": ["Transaction Amount"],
    },
]


def _normalize_text(value):
    return " ".join(str(value or "").strip().lower().split())


def _canonical_term(value):
    return _normalize_text(value).replace("_", " ")


def _compact_list(values):
    return [value for value in values if value]


def _append_unique(target_list, values):
    seen = set(target_list)
    for value in values:
        if value and value not in seen:
            target_list.append(value)
            seen.add(value)


def _build_entry(term, term_type, definition, business_usage, status, synonyms=None, notes=None):
    return {
        "term": term,
        "term_type": term_type,
        "definition": definition.strip(),
        "business_usage": business_usage.strip(),
        "synonyms": _compact_list(synonyms or []),
        "notes": notes or "",
        "status": status,
    }


def _register_entry(bucket, entry):
    key = _canonical_term(entry["term"])
    if not key:
        return
    if key not in bucket:
        bucket[key] = dict(entry)
        return

    existing = bucket[key]
    if existing.get("status") == "draft" and entry.get("status") in {"approved", "rejected"}:
        for field in ["definition", "business_usage", "notes", "status"]:
            existing[field] = entry[field]
    _append_unique(existing["synonyms"], entry["synonyms"])


def load_brief(input_path):
    input_path = Path(input_path)
    raw = input_path.read_text(encoding="utf-8").strip()
    if input_path.suffix.lower() == ".json":
        payload = json.loads(raw)
        if not isinstance(payload, dict):
            raise ValueError("Business brief JSON must be an object.")
        if "tables" in payload or "concept_registry" in payload:
            raise ValueError("Schema-shaped JSON is no longer supported. Provide a business domain description instead.")
        return {
            "domain": str(payload.get("domain") or "").strip(),
            "description": str(payload.get("description") or "").strip(),
            "core_processes": [str(item).strip() for item in payload.get("core_processes") or [] if str(item).strip()],
            "entities": [str(item).strip() for item in payload.get("entities") or [] if str(item).strip()],
            "notes": str(payload.get("notes") or "").strip(),
        }
    return {
        "domain": "",
        "description": raw,
        "core_processes": [],
        "entities": [],
        "notes": "",
    }


def _combined_text(brief):
    return _normalize_text(
        " ".join(
            [
                brief.get("domain", ""),
                brief.get("description", ""),
                " ".join(brief.get("core_processes") or []),
                " ".join(brief.get("entities") or []),
                brief.get("notes", ""),
            ]
        )
    )


def _process_entries(brief):
    entries = []
    for process in brief.get("core_processes") or []:
        title = " ".join(word.capitalize() for word in process.replace("-", " ").split())
        entries.append(
            _build_entry(
                term=title,
                term_type="business_process",
                definition=f"A core business process within the {brief.get('domain') or 'described'} operating model.",
                business_usage="Used to organize the operating lifecycle and the main flow of work described in the brief.",
                status="draft",
                synonyms=[process],
            )
        )
    return entries


def _entity_entries(brief):
    entries = []
    for entity in brief.get("entities") or []:
        title = " ".join(word.capitalize() for word in entity.replace("-", " ").split())
        entries.append(
            _build_entry(
                term=title,
                term_type="business_entity",
                definition=f"A core business entity within the {brief.get('domain') or 'described'} operating model.",
                business_usage="Used as a primary business concept in the domain brief and related operating workflows.",
                status="draft",
                synonyms=[entity],
            )
        )
    return entries


def _keyword_entries(brief):
    text = _combined_text(brief)
    entries = []
    for term_config in KEYWORD_LIBRARY:
        if not any(keyword in text for keyword in term_config["keywords"]):
            continue
        entries.append(
            _build_entry(
                term=term_config["term"],
                term_type=term_config["term_type"],
                definition=term_config["definition"],
                business_usage=term_config["business_usage"],
                status="draft",
                synonyms=term_config["synonyms"],
            )
        )
    return entries


def build_glossary(brief, source_path):
    bucket = {}
    for entry in _process_entries(brief):
        _register_entry(bucket, entry)
    for entry in _entity_entries(brief):
        _register_entry(bucket, entry)
    for entry in _keyword_entries(brief):
        _register_entry(bucket, entry)

    entries = sorted(bucket.values(), key=lambda item: (item["term"].lower(), item["term_type"]))
    metadata = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_description_path": str(Path(source_path)),
        "glossary_entry_count": len(entries),
        "generation_mode": "domain_brief",
        "inference_mode": "domain_guided",
    }
    return {"metadata": metadata, "entries": entries}


def _autofit(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _apply_status_dropdown(ws, headers):
    status_col_idx = headers.index("status") + 1
    status_col_letter = ws.cell(row=1, column=status_col_idx).column_letter
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_VALUES) + '"',
        allow_blank=True,
    )
    validation.prompt = "Select glossary review status"
    validation.error = "Choose one of: draft, approved, rejected."
    ws.add_data_validation(validation)
    validation.add(f"{status_col_letter}2:{status_col_letter}1048576")


def write_glossary_excel(glossary, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    headers = [
        "term",
        "term_type",
        "definition",
        "business_usage",
        "synonyms",
        "notes",
        "status",
    ]
    ws.append(headers)
    for row in glossary.get("entries") or []:
        ws.append(
            [
                row.get("term", ""),
                row.get("term_type", ""),
                row.get("definition", ""),
                row.get("business_usage", ""),
                ", ".join(row.get("synonyms") or []),
                row.get("notes", ""),
                row.get("status", ""),
            ]
        )

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _apply_status_dropdown(ws, headers)
    _autofit(ws)

    meta_ws = wb.create_sheet("RunMetadata")
    meta_ws.append(["Field", "Value"])
    for cell in meta_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for key, value in glossary.get("metadata", {}).items():
        meta_ws.append([key, value])
    meta_ws.freeze_panes = "A2"
    meta_ws.auto_filter.ref = meta_ws.dimensions
    _autofit(meta_ws)

    wb.save(output_path)


def output_paths(input_path):
    input_path = Path(input_path)
    stem = input_path.stem
    return (
        input_path.with_name(f"{stem}_glossary.json"),
        input_path.with_name(f"{stem}_glossary.xlsx"),
    )


def run(input_path):
    source_path = Path(input_path)
    brief = load_brief(source_path)
    glossary = build_glossary(brief, source_path)
    json_output, excel_output = output_paths(source_path)
    json_output.write_text(json.dumps(glossary, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    write_glossary_excel(glossary, excel_output)
    return json_output, excel_output, glossary


def main():
    parser = argparse.ArgumentParser(description="Generate glossary JSON and Excel from a business domain description.")
    parser.add_argument("input_brief", help="Path to a text, markdown, or JSON business brief.")
    args = parser.parse_args()
    json_output, excel_output, glossary = run(args.input_brief)
    print(
        json.dumps(
            {
                "json_output": str(json_output),
                "excel_output": str(excel_output),
                "glossary_entry_count": glossary["metadata"]["glossary_entry_count"],
            },
            ensure_ascii=True,
        )
    )


if __name__ == "__main__":
    main()
