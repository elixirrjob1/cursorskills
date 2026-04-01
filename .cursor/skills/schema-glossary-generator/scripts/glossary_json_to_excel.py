#!/usr/bin/env python3
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


EXPECTED_HEADERS = [
    "term",
    "term_type",
    "definition",
    "business_usage",
    "synonyms",
    "source_tables",
    "source_columns",
    "confidence",
    "confidence_tier",
    "inference_basis",
    "source_refs",
    "notes",
    "status",
]

HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _list_to_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def normalize_payload(payload):
    if isinstance(payload, list):
        payload = {"metadata": {}, "entries": payload}
    if not isinstance(payload, dict):
        raise ValueError("Glossary payload must be a JSON object or a list of entries.")

    metadata = payload.get("metadata") or {}
    entries = payload.get("entries") or []
    if not isinstance(entries, list):
        raise ValueError("Glossary payload 'entries' must be a list.")

    normalized_entries = []
    for index, entry in enumerate(entries, start=1):
        if not isinstance(entry, dict):
            raise ValueError(f"Glossary entry {index} must be an object.")
        normalized = {header: entry.get(header, "" if header not in {"synonyms", "source_tables", "source_columns", "source_refs"} else []) for header in EXPECTED_HEADERS}
        normalized_entries.append(normalized)
    return {"metadata": metadata, "entries": normalized_entries}


def _autofit(ws):
    for column_cells in ws.columns:
        max_length = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def write_workbook(payload, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"

    ws.append(EXPECTED_HEADERS)
    for entry in payload["entries"]:
        ws.append(
            [
                entry["term"],
                entry["term_type"],
                entry["definition"],
                entry["business_usage"],
                _list_to_text(entry["synonyms"]),
                _list_to_text(entry["source_tables"]),
                _list_to_text(entry["source_columns"]),
                entry["confidence"],
                entry["confidence_tier"],
                entry["inference_basis"],
                _list_to_text(entry["source_refs"]),
                entry["notes"],
                entry["status"],
            ]
        )

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws)

    meta_ws = wb.create_sheet("RunMetadata")
    meta_ws.append(["Field", "Value"])
    for cell in meta_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for key, value in payload["metadata"].items():
        meta_ws.append([key, _list_to_text(value)])
    meta_ws.freeze_panes = "A2"
    meta_ws.auto_filter.ref = meta_ws.dimensions
    _autofit(meta_ws)

    wb.save(output_path)


def default_output_path(input_path):
    input_path = Path(input_path)
    return input_path.with_suffix(".xlsx")


def run(input_json, output_xlsx=None):
    input_path = Path(input_json)
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    normalized = normalize_payload(payload)
    output_path = Path(output_xlsx) if output_xlsx else default_output_path(input_path)
    write_workbook(normalized, output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Export glossary JSON to Excel.")
    parser.add_argument("input_json", help="Path to glossary JSON written by the agent.")
    parser.add_argument("output_xlsx", nargs="?", help="Optional output workbook path.")
    args = parser.parse_args()
    output_path = run(args.input_json, args.output_xlsx)
    print(json.dumps({"excel_output": str(output_path)}, ensure_ascii=True))


if __name__ == "__main__":
    main()
