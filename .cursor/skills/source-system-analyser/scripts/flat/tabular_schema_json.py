#!/usr/bin/env python3
"""Convert CSV/XLSX tabular schema metadata to schema.json-like structure and back."""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    import yaml  # type: ignore
except Exception:
    yaml = None


def _norm_header(name: str) -> str:
    return str(name or "").strip().lower().replace(" ", "_")


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in {"1", "true", "t", "yes", "y"}


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except Exception:
        return None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _validate_delimiter(delimiter: str | None) -> str | None:
    if delimiter is None:
        return None
    if len(delimiter) != 1:
        raise SystemExit("Delimiter must be a single character (for example ',', ';', '|', or '\\t').")
    return delimiter


def _detect_csv_delimiter(path: Path, fallback: str = ",") -> str:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
    if not sample:
        return fallback
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except Exception:
        return fallback


def _resolve_csv_delimiter(path: Path, delimiter: str | None) -> str:
    validated = _validate_delimiter(delimiter)
    if validated is not None:
        return validated
    return _detect_csv_delimiter(path)


def _read_csv(path: Path, delimiter: str | None = None) -> list[dict[str, Any]]:
    resolved = _resolve_csv_delimiter(path, delimiter)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=resolved)
        out: list[dict[str, Any]] = []
        for row in reader:
            out.append({_norm_header(k): v for k, v in row.items()})
        return out


def _read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Reading .xlsx requires openpyxl. Install with: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers_raw = next(rows, None)
    if not headers_raw:
        return []
    headers = [_norm_header(h) for h in headers_raw]
    data: list[dict[str, Any]] = []
    for r in rows:
        if r is None:
            continue
        item = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        if any(v not in (None, "") for v in item.values()):
            data.append(item)
    return data


def _read_tabular(path: str, sheet_name: str | None = None, delimiter: str | None = None) -> list[dict[str, Any]]:
    p = Path(path)
    if p.suffix.lower() == ".csv":
        return _read_csv(p, delimiter=delimiter)
    if p.suffix.lower() == ".xlsx":
        return _read_xlsx(p, sheet_name)
    raise SystemExit(f"Unsupported file type: {p.suffix}. Use .csv or .xlsx")


def _get(row: dict[str, Any], *keys: str, default: Any = None) -> Any:
    for k in keys:
        if k in row and row.get(k) not in (None, ""):
            return row.get(k)
    return default


def _headers(rows: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                out.append(k)
    return out


def _fk_to_join_candidate(fk: dict[str, Any]) -> dict[str, Any] | None:
    column = fk.get("column")
    references = fk.get("references")
    if not column:
        return None
    target_table = None
    target_column = None
    if isinstance(references, str) and "." in references:
        target_table, target_column = references.split(".", 1)
    return {
        "column": column,
        "target_table": target_table,
        "target_column": target_column,
        "confidence": "high",
    }


_SEMANTIC_PATTERNS: list[tuple[str, str]] = [
    (r"(length|height|width|depth|distance|diameter|radius|thickness)", "length"),
    (r"(volume|capacity|cubic|cbm|ft3|m3|liter|litre|gallon)", "volume"),
    (r"(pressure|press|psi|bar|kpa|mpa)", "pressure"),
    (r"(temperature|temp|celsius|fahrenheit|kelvin)", "temperature"),
    (r"(duration|latency|elapsed|runtime|ttl|age|timeout)", "duration"),
]
_UNITFUL_SEMANTICS = {"length", "area", "volume", "mass", "pressure", "temperature", "duration", "speed", "flow_rate", "force", "energy", "power", "density"}
_UNIT_ALIASES = {
    "ft": "ft",
    "feet": "ft",
    "foot": "ft",
    "in": "in",
    "inch": "in",
    "inches": "in",
    "m": "m",
    "meter": "m",
    "meters": "m",
    "cm": "cm",
    "mm": "mm",
    "ft3": "ft3",
    "cubic_ft": "ft3",
    "m3": "m3",
    "cubic_m": "m3",
    "psi": "psi",
    "bar": "bar",
    "kpa": "kpa",
    "mpa": "mpa",
    "s": "s",
    "sec": "s",
    "second": "s",
    "min": "min",
    "minute": "min",
    "h": "h",
    "hour": "h",
    "c": "c",
    "celsius": "c",
    "f": "f",
    "fahrenheit": "f",
    "k": "k",
    "kelvin": "k",
}
_UNIT_CONVERSION: dict[str, dict[str, Any]] = {
    "ft": {"canonical_unit": "m", "unit_system": "imperial", "factor_to_canonical": 0.3048, "offset_to_canonical": 0.0},
    "in": {"canonical_unit": "m", "unit_system": "imperial", "factor_to_canonical": 0.0254, "offset_to_canonical": 0.0},
    "m": {"canonical_unit": "m", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": 0.0},
    "cm": {"canonical_unit": "m", "unit_system": "metric", "factor_to_canonical": 0.01, "offset_to_canonical": 0.0},
    "mm": {"canonical_unit": "m", "unit_system": "metric", "factor_to_canonical": 0.001, "offset_to_canonical": 0.0},
    "ft3": {"canonical_unit": "m3", "unit_system": "imperial", "factor_to_canonical": 0.028316846592, "offset_to_canonical": 0.0},
    "m3": {"canonical_unit": "m3", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": 0.0},
    "psi": {"canonical_unit": "bar", "unit_system": "imperial", "factor_to_canonical": 0.0689475729, "offset_to_canonical": 0.0},
    "bar": {"canonical_unit": "bar", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": 0.0},
    "kpa": {"canonical_unit": "bar", "unit_system": "metric", "factor_to_canonical": 0.01, "offset_to_canonical": 0.0},
    "mpa": {"canonical_unit": "bar", "unit_system": "metric", "factor_to_canonical": 10.0, "offset_to_canonical": 0.0},
    "s": {"canonical_unit": "s", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": 0.0},
    "min": {"canonical_unit": "s", "unit_system": "metric", "factor_to_canonical": 60.0, "offset_to_canonical": 0.0},
    "h": {"canonical_unit": "s", "unit_system": "metric", "factor_to_canonical": 3600.0, "offset_to_canonical": 0.0},
    "c": {"canonical_unit": "c", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": 0.0},
    "f": {"canonical_unit": "c", "unit_system": "imperial", "factor_to_canonical": 0.5555555556, "offset_to_canonical": -17.7777777778},
    "k": {"canonical_unit": "c", "unit_system": "metric", "factor_to_canonical": 1.0, "offset_to_canonical": -273.15},
}
_RULES_LOADED = False


def _load_context_rules() -> None:
    global _RULES_LOADED, _SEMANTIC_PATTERNS, _UNIT_ALIASES, _UNIT_CONVERSION
    if _RULES_LOADED:
        return
    _RULES_LOADED = True
    if yaml is None:
        return
    shared_dir = (Path(__file__).resolve().parent.parent / "references" / "shared")
    semantic_path = shared_dir / "semantic_mappings.yaml"
    unit_path = shared_dir / "unit_mappings.yaml"
    try:
        if semantic_path.exists():
            with open(semantic_path, encoding="utf-8") as f:
                data = yaml.safe_load(f) or {}
            patterns = data.get("semantic_patterns")
            if isinstance(patterns, list):
                parsed: list[tuple[str, str]] = []
                for item in patterns:
                    if not isinstance(item, dict):
                        continue
                    pat = str(item.get("pattern") or "").strip()
                    cls = str(item.get("semantic_class") or "").strip()
                    if pat and cls:
                        parsed.append((pat, cls))
                if parsed:
                    _SEMANTIC_PATTERNS = parsed
    except Exception:
        pass
    try:
        if unit_path.exists():
            with open(unit_path, encoding="utf-8") as f:
                data = yaml.safe_load(f) or {}
            aliases = data.get("unit_aliases")
            if isinstance(aliases, dict):
                _UNIT_ALIASES.update({str(k): str(v) for k, v in aliases.items()})
            conversions = data.get("unit_conversion")
            if isinstance(conversions, dict):
                _UNIT_CONVERSION.update(conversions)
    except Exception:
        pass


def _infer_semantic_class(col_name: str) -> str | None:
    _load_context_rules()
    lower = col_name.lower()
    for pattern, semantic_class in _SEMANTIC_PATTERNS:
        if re.search(pattern, lower):
            return semantic_class
    return None


def _extract_unit_from_name(col_name: str) -> str | None:
    _load_context_rules()
    lower = re.sub(r"[^a-z0-9]+", "_", col_name.lower()).strip("_")
    if not lower:
        return None
    for alias in sorted(_UNIT_ALIASES.keys(), key=len, reverse=True):
        norm_alias = re.sub(r"[^a-z0-9]+", "_", alias).strip("_")
        if re.search(rf"(?:^|_){re.escape(norm_alias)}(?:$|_)", lower):
            return _UNIT_ALIASES[alias]
    return None


def _build_unit_context(col_name: str, semantic_class: str | None, raw_row: dict[str, Any]) -> dict[str, Any] | None:
    _load_context_rules()
    detected = _get(raw_row, "detected_unit", "unit")
    if detected not in (None, ""):
        detected = str(detected).strip().lower()
    else:
        detected = _extract_unit_from_name(col_name)
    canonical = _get(raw_row, "canonical_unit")
    if canonical in (None, "") and detected:
        canonical = _UNIT_CONVERSION.get(detected, {}).get("canonical_unit")
    unit_system = _get(raw_row, "unit_system")
    if unit_system in (None, "") and detected:
        unit_system = _UNIT_CONVERSION.get(detected, {}).get("unit_system", "unknown")
    factor = _as_float(_get(raw_row, "factor_to_canonical"))
    offset = _as_float(_get(raw_row, "offset_to_canonical"))
    if factor is None and detected in _UNIT_CONVERSION:
        factor = _UNIT_CONVERSION[detected].get("factor_to_canonical")
    if offset is None and detected in _UNIT_CONVERSION:
        offset = _UNIT_CONVERSION[detected].get("offset_to_canonical")

    if not detected:
        if semantic_class in _UNITFUL_SEMANTICS:
            return {
                "detected_unit": None,
                "canonical_unit": None,
                "unit_system": "unknown",
                "conversion": None,
                "detection_confidence": "low",
                "detection_source": "combined",
                "notes": "Semantic class suggests units, but no explicit source unit token was detected.",
            }
        return None

    conversion = None
    if factor is not None:
        offset_val = 0.0 if offset is None else offset
        conversion = {
            "factor_to_canonical": factor,
            "offset_to_canonical": offset_val,
            "formula": f"canonical = value * {factor} + {offset_val}",
        }

    return {
        "detected_unit": detected,
        "canonical_unit": canonical,
        "unit_system": unit_system or "unknown",
        "conversion": conversion,
        "detection_confidence": "medium",
        "detection_source": "combined",
        "notes": None if not canonical or detected == canonical else f"Normalize from '{detected}' to canonical '{canonical}'.",
    }


def _build_unit_summary(columns: list[dict[str, Any]]) -> dict[str, Any]:
    with_units = 0
    unknown_cols: list[str] = []
    groups: dict[str, set[str]] = {}
    for col in columns:
        semantic_class = col.get("semantic_class")
        unit_ctx = col.get("unit_context")
        if isinstance(unit_ctx, dict) and unit_ctx.get("detected_unit"):
            with_units += 1
            if semantic_class:
                groups.setdefault(str(semantic_class), set()).add(str(unit_ctx.get("detected_unit")))
        elif semantic_class in _UNITFUL_SEMANTICS:
            unknown_cols.append(str(col.get("name")))
    mixed_groups = [
        {"semantic_class": semantic_class, "detected_units": sorted(units)}
        for semantic_class, units in groups.items()
        if len(units) > 1
    ]
    return {
        "columns_with_units": with_units,
        "columns_without_units": max(len(columns) - with_units, 0),
        "mixed_unit_groups": mixed_groups,
        "unknown_unit_columns": sorted(unknown_cols),
    }


def _suggest_role(headers: list[str]) -> dict[str, str | None]:
    s = set(headers)

    def pick(*opts: str) -> str | None:
        for o in opts:
            if o in s:
                return o
        return None

    return {
        "table": pick("table", "table_name", "entity", "object"),
        "column": pick("column", "column_name", "name", "field", "attribute"),
        "type": pick("type", "data_type", "dtype", "column_type"),
        "schema": pick("schema", "schema_name"),
        "primary_key": pick("primary_key", "pk"),
        "foreign_key": pick("foreign_key", "fk"),
    }


def _parse_columns(rows: list[dict[str, Any]], default_schema: str) -> dict[str, dict[str, Any]]:
    tables: dict[str, dict[str, Any]] = {}
    for r in rows:
        table = str(_get(r, "table", "table_name", "entity", "object", default="")).strip()
        column = str(_get(r, "column", "column_name", "name", "field", "attribute", default="")).strip()
        col_type = str(_get(r, "type", "data_type", "dtype", "column_type", default="text")).strip() or "text"
        if not table or not column:
            continue
        schema = str(_get(r, "schema", "schema_name", default=default_schema)).strip() or default_schema

        t = tables.setdefault(
            table,
            {
                "table": table,
                "schema": schema,
                "table_description": None,
                "columns": [],
                "primary_keys": [],
                "foreign_keys": [],
                "row_count": _as_int(_get(r, "row_count")) or 0,
                "field_classifications": [],
                "sensitive_fields": [],
                "incremental_columns": [],
                "partition_columns": [],
                "join_candidates": [],
                "unit_summary": {
                    "columns_with_units": 0,
                    "columns_without_units": 0,
                    "mixed_unit_groups": [],
                    "unknown_unit_columns": [],
                },
                "cdc_enabled": _as_bool(_get(r, "cdc_enabled"), False),
                "has_primary_key": False,
                "has_foreign_keys": False,
                "has_sensitive_fields": False,
                "data_quality": {
                    "findings": [],
                    "summary": {"critical": 0, "warning": 0, "info": 0},
                    "constraints_found": {},
                },
            },
        )

        is_incremental = _as_bool(_get(r, "is_incremental", "incremental"), False)
        semantic_class = _get(r, "semantic_class")
        if semantic_class in (None, ""):
            semantic_class = _infer_semantic_class(column)
        else:
            semantic_class = str(semantic_class).strip()
        col = {
            "name": column,
            "type": col_type,
            "nullable": _as_bool(_get(r, "nullable", "is_nullable"), True),
            "column_description": _get(r, "column_description", "description"),
            "is_incremental": is_incremental,
            "cardinality": _as_int(_get(r, "cardinality", "distinct_count")),
            "null_count": _as_int(_get(r, "null_count", "nulls")),
            "data_range": {
                "min": None if _get(r, "min", "min_value") in (None, "") else str(_get(r, "min", "min_value")),
                "max": None if _get(r, "max", "max_value") in (None, "") else str(_get(r, "max", "max_value")),
            },
            "data_category": _get(r, "data_category", "category") or None,
            "semantic_class": semantic_class,
            "unit_context": _build_unit_context(column, semantic_class, r),
        }
        t["columns"].append(col)

        if _as_bool(_get(r, "primary_key", "pk"), False):
            t["primary_keys"].append(column)
        if is_incremental:
            t["incremental_columns"].append(column)

        fk = str(_get(r, "foreign_key", "fk", default="")).strip()
        if fk:
            t["foreign_keys"].append({"column": column, "references": fk})

    for t in tables.values():
        t["primary_keys"] = sorted(set(t["primary_keys"]))
        t["incremental_columns"] = sorted(set(t["incremental_columns"]))
        # dedupe foreign keys by (column, references)
        seen = set()
        fks = []
        for fk in t["foreign_keys"]:
            key = (fk.get("column"), fk.get("references"))
            if key not in seen:
                seen.add(key)
                fks.append(fk)
        t["foreign_keys"] = fks
        t["join_candidates"] = [jc for fk in fks if (jc := _fk_to_join_candidate(fk)) is not None]
        t["unit_summary"] = _build_unit_summary(t["columns"])
        t["has_primary_key"] = bool(t["primary_keys"])
        t["has_foreign_keys"] = bool(t["foreign_keys"])
    return tables


def _merge_table_rows(tables: dict[str, dict[str, Any]], rows: list[dict[str, Any]], default_schema: str) -> None:
    for r in rows:
        table = str(_get(r, "table", "table_name", "entity", "object", default="")).strip()
        if not table:
            continue
        t = tables.setdefault(
            table,
            {
                "table": table,
                "schema": str(_get(r, "schema", "schema_name", default=default_schema)).strip() or default_schema,
                "table_description": None,
                "columns": [],
                "primary_keys": [],
                "foreign_keys": [],
                "row_count": 0,
                "field_classifications": [],
                "sensitive_fields": [],
                "incremental_columns": [],
                "partition_columns": [],
                "join_candidates": [],
                "unit_summary": {
                    "columns_with_units": 0,
                    "columns_without_units": 0,
                    "mixed_unit_groups": [],
                    "unknown_unit_columns": [],
                },
                "cdc_enabled": False,
                "has_primary_key": False,
                "has_foreign_keys": False,
                "has_sensitive_fields": False,
                "data_quality": {
                    "findings": [],
                    "summary": {"critical": 0, "warning": 0, "info": 0},
                    "constraints_found": {},
                },
            },
        )
        if _get(r, "schema", "schema_name") not in (None, ""):
            t["schema"] = str(_get(r, "schema", "schema_name")).strip()
        if _get(r, "table_description", "description") not in (None, ""):
            t["table_description"] = str(_get(r, "table_description", "description")).strip()
        if _as_int(_get(r, "row_count", "rows")) is not None:
            t["row_count"] = _as_int(_get(r, "row_count", "rows")) or 0
        if _get(r, "cdc_enabled", "cdc") not in (None, ""):
            t["cdc_enabled"] = _as_bool(_get(r, "cdc_enabled", "cdc"), False)


def cmd_inspect(args: argparse.Namespace) -> None:
    out: dict[str, Any] = {"files": []}

    def add_file(path: str, sheet: str | None, kind: str, delimiter: str | None = None) -> None:
        rows = _read_tabular(path, sheet, delimiter=delimiter)
        p = Path(path)
        headers = _headers(rows)
        entry = {
            "kind": kind,
            "path": path,
            "sheet": sheet,
            "row_count": len(rows),
            "headers": headers,
            "suggested_mapping": _suggest_role(headers),
            "sample_rows": rows[: args.sample_size],
        }
        if p.suffix.lower() == ".csv":
            entry["delimiter"] = _resolve_csv_delimiter(p, delimiter)
        out["files"].append(entry)

    add_file(args.columns_file, args.columns_sheet, "columns", delimiter=args.columns_delimiter)
    if args.tables_file:
        add_file(args.tables_file, args.tables_sheet, "tables", delimiter=args.tables_delimiter)

    payload = json.dumps(out, indent=2, default=str)
    if args.output:
        Path(args.output).write_text(payload, encoding="utf-8")
        print(f"Wrote {args.output}")
    else:
        print(payload)


def cmd_to_json(args: argparse.Namespace) -> None:
    col_rows = _read_tabular(args.columns_file, args.columns_sheet, delimiter=args.columns_delimiter)
    tables = _parse_columns(col_rows, args.default_schema)

    if args.tables_file:
        table_rows = _read_tabular(args.tables_file, args.tables_sheet, delimiter=args.tables_delimiter)
        _merge_table_rows(tables, table_rows, args.default_schema)

    ordered_tables = [tables[k] for k in sorted(tables.keys())]
    out = {
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "database_url": None,
            "schema_filter": args.default_schema,
            "total_tables": len(ordered_tables),
            "total_rows": sum(int(t.get("row_count") or 0) for t in ordered_tables),
            "total_findings": 0,
        },
        "connection": {
            "host": None,
            "port": None,
            "database": None,
            "driver": None,
            "timezone": None,
        },
        "source_system_context": {
            "contacts": [],
            "delete_management_instruction": "",
            "restrictions": "",
            "late_arriving_data_manual": "",
            "volume_size_projection_manual": "",
            "field_context_manual": "",
        },
        "data_quality_summary": {
            "critical": 0,
            "warning": 0,
            "info": 0,
            "by_check": {},
            "constraints_found": {},
        },
        "tables": ordered_tables,
    }

    Path(args.output).write_text(json.dumps(out, indent=2), encoding="utf-8")
    print(f"Wrote {args.output} (tables={len(ordered_tables)})")


def cmd_from_json(args: argparse.Namespace) -> None:
    src = json.loads(Path(args.schema_json).read_text(encoding="utf-8"))
    tables = src.get("tables", [])

    col_headers = [
        "table",
        "schema",
        "column",
        "column_description",
        "type",
        "nullable",
        "is_incremental",
        "primary_key",
        "foreign_key",
        "cardinality",
        "null_count",
        "min",
        "max",
        "data_category",
        "semantic_class",
        "detected_unit",
        "canonical_unit",
        "unit_system",
        "factor_to_canonical",
        "offset_to_canonical",
    ]
    table_headers = ["table", "schema", "table_description", "row_count", "cdc_enabled"]

    col_rows: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []

    for t in tables:
        tname = t.get("table")
        schema = t.get("schema")
        pk_set = set(t.get("primary_keys") or [])
        fk_map = {fk.get("column"): fk.get("references") for fk in (t.get("foreign_keys") or [])}

        table_rows.append(
            {
                "table": tname,
                "schema": schema,
                "table_description": t.get("table_description"),
                "row_count": t.get("row_count"),
                "cdc_enabled": t.get("cdc_enabled"),
            }
        )

        for c in t.get("columns") or []:
            rng = c.get("data_range") or {}
            unit_ctx = c.get("unit_context") if isinstance(c.get("unit_context"), dict) else {}
            conversion = unit_ctx.get("conversion") if isinstance(unit_ctx.get("conversion"), dict) else {}
            col_rows.append(
                {
                    "table": tname,
                    "schema": schema,
                    "column": c.get("name"),
                    "column_description": c.get("column_description"),
                    "type": c.get("type"),
                    "nullable": c.get("nullable"),
                    "is_incremental": c.get("is_incremental"),
                    "primary_key": c.get("name") in pk_set,
                    "foreign_key": fk_map.get(c.get("name"), ""),
                    "cardinality": c.get("cardinality"),
                    "null_count": c.get("null_count"),
                    "min": rng.get("min"),
                    "max": rng.get("max"),
                    "data_category": c.get("data_category"),
                    "semantic_class": c.get("semantic_class"),
                    "detected_unit": unit_ctx.get("detected_unit"),
                    "canonical_unit": unit_ctx.get("canonical_unit"),
                    "unit_system": unit_ctx.get("unit_system"),
                    "factor_to_canonical": conversion.get("factor_to_canonical"),
                    "offset_to_canonical": conversion.get("offset_to_canonical"),
                }
            )

    with Path(args.columns_out).open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=col_headers)
        w.writeheader()
        w.writerows(col_rows)

    with Path(args.tables_out).open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=table_headers)
        w.writeheader()
        w.writerows(table_rows)

    print(f"Wrote {args.columns_out} and {args.tables_out}")


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    p_inspect = sub.add_parser("inspect", help="Inspect tabular file columns/rows before conversion")
    p_inspect.add_argument("--columns-file", required=True, help="Path to CSV/XLSX with column-level metadata")
    p_inspect.add_argument("--columns-sheet", default=None, help="Sheet name when --columns-file is .xlsx")
    p_inspect.add_argument("--columns-delimiter", default=None, help="CSV delimiter for --columns-file (auto-detected when omitted)")
    p_inspect.add_argument("--tables-file", default=None, help="Optional CSV/XLSX with table-level metadata")
    p_inspect.add_argument("--tables-sheet", default=None, help="Sheet name when --tables-file is .xlsx")
    p_inspect.add_argument("--tables-delimiter", default=None, help="CSV delimiter for --tables-file (auto-detected when omitted)")
    p_inspect.add_argument("--sample-size", type=int, default=5, help="Number of sample rows per file")
    p_inspect.add_argument("--output", default=None, help="Optional output JSON path")
    p_inspect.set_defaults(func=cmd_inspect)

    p_to = sub.add_parser("to-json", help="Convert tabular schema files to schema.json-like output")
    p_to.add_argument("--columns-file", required=True, help="Path to CSV/XLSX with column-level metadata")
    p_to.add_argument("--columns-sheet", default=None, help="Sheet name when --columns-file is .xlsx")
    p_to.add_argument("--columns-delimiter", default=None, help="CSV delimiter for --columns-file (auto-detected when omitted)")
    p_to.add_argument("--tables-file", default=None, help="Optional CSV/XLSX with table-level metadata")
    p_to.add_argument("--tables-sheet", default=None, help="Sheet name when --tables-file is .xlsx")
    p_to.add_argument("--tables-delimiter", default=None, help="CSV delimiter for --tables-file (auto-detected when omitted)")
    p_to.add_argument("--output", required=True, help="Output JSON path")
    p_to.add_argument("--default-schema", default="public", help="Schema name used when missing in source")
    p_to.set_defaults(func=cmd_to_json)

    p_from = sub.add_parser("from-json", help="Export schema.json to CSV templates")
    p_from.add_argument("--schema-json", required=True, help="Input schema.json path")
    p_from.add_argument("--columns-out", required=True, help="Output columns CSV path")
    p_from.add_argument("--tables-out", required=True, help="Output tables CSV path")
    p_from.set_defaults(func=cmd_from_json)

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
