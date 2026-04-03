import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path("/home/fillip/projec/cursorskills/.cursor/skills/schema-glossary-generator/scripts/glossary_json_to_excel.py")


def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


export_module = _load_module("glossary_json_to_excel", SCRIPT_PATH)


class GlossaryJsonToExcelTests(unittest.TestCase):
    def test_normalize_payload_fills_expected_columns(self):
        payload = {
            "metadata": {"source_description_path": "domain.txt"},
            "entries": [
                {
                    "term": "Customer",
                    "term_type": "business_entity",
                    "definition": "Customer master record.",
                    "business_usage": "Used in sales reporting.",
                }
            ],
        }

        normalized = export_module.normalize_payload(payload)
        entry = normalized["entries"][0]
        self.assertEqual(list(entry.keys()), export_module.EXPECTED_HEADERS)
        self.assertEqual(entry["synonyms"], [])
        self.assertEqual(entry["status"], "")

    def test_run_writes_expected_workbook(self):
        payload = {
            "metadata": {
                "source_description_path": "domain.txt",
                "generation_mode": "agent_authored",
            },
            "entries": [
                {
                    "term": "Sales Order",
                    "term_type": "business_process",
                    "definition": "Customer sales transaction header.",
                    "business_usage": "Used to track sales lifecycle and order totals.",
                    "synonyms": ["Order"],
                    "notes": "",
                    "status": "draft",
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "glossary.json"
            input_path.write_text(json.dumps(payload), encoding="utf-8")
            output_path = export_module.run(input_path)
            wb = load_workbook(output_path)
            self.assertEqual(wb.sheetnames, ["Glossary", "RunMetadata"])
            headers = [cell.value for cell in wb["Glossary"][1]]
            self.assertEqual(headers, export_module.EXPECTED_HEADERS)
            self.assertEqual(wb["Glossary"]["A2"].value, "Sales Order")
            validations = list(wb["Glossary"].data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].formula1, '"draft,approved,rejected"')
            self.assertIn("G2", str(validations[0].sqref))


if __name__ == "__main__":
    unittest.main()
