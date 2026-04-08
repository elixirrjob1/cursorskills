import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


WRITER_PATH = Path("/home/fillip/projec/cursorskills/.cursor/skills/json-to-excel-export/scripts/json_to_excel.py")
READER_PATH = Path("/home/fillip/projec/cursorskills/.cursor/skills/json-to-excel-export/scripts/excel_to_json.py")


def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


json_to_excel = _load_module("json_to_excel", WRITER_PATH)
excel_to_json = _load_module("excel_to_json", READER_PATH)


def _section_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    sections = {}
    i = 0
    while i < len(rows):
        row = rows[i] or ()
        name = row[0] if row else None
        if name in (None, ""):
            i += 1
            continue
        headers = [h for h in (rows[i + 1] or ()) if h not in (None, "")]
        i += 2
        data_rows = []
        while i < len(rows):
            current = rows[i] or ()
            if all(cell in (None, "") for cell in current):
                i += 1
                break
            item = {}
            for idx, header in enumerate(headers):
                item[str(header)] = current[idx] if idx < len(current) else None
            data_rows.append(item)
            i += 1
        sections[str(name)] = data_rows
    return sections


def _set_section_value(ws, section_name, target_col, value, match_col=None, match_value=None):
    rows = list(ws.iter_rows())
    for idx, row in enumerate(rows):
        if not row:
            continue
        if row[0].value != section_name:
            continue
        headers = [cell.value for cell in rows[idx + 1]]
        target_idx = headers.index(target_col)
        data_idx = idx + 2
        while data_idx < len(rows):
            current = rows[data_idx]
            if all(cell.value in (None, "") for cell in current):
                break
            if match_col is None:
                current[target_idx].value = value
                return
            match_idx = headers.index(match_col)
            if current[match_idx].value == match_value:
                current[target_idx].value = value
                return
            data_idx += 1
    raise AssertionError(f"Could not find section={section_name} target={target_col} match={match_col}:{match_value}")


class JsonExcelLayoutTests(unittest.TestCase):
    def setUp(self):
        self.payload = {
            "metadata": {
                "generated_at": "2026-03-09T12:00:00Z",
                "database_url": "postgresql://user:pass@localhost:5432/demo",
                "schema_filter": "public",
                "total_tables": 2,
                "total_rows": 15,
                "total_findings": 1,
                "openmetadata_classifications": [
                    {
                        "name": "PII",
                        "provider": "system",
                        "description": "Personally identifiable information.",
                        "mutually_exclusive": True,
                        "allowed_on": ["table", "column"],
                        "options": [
                            {"name": "Sensitive", "fqn": "PII.Sensitive", "description": "Sensitive personal data."},
                            {"name": "NonSensitive", "fqn": "PII.NonSensitive", "description": "Non-sensitive personal data."},
                        ],
                    },
                    {
                        "name": "Tier",
                        "provider": "system",
                        "description": "Business criticality tiers.",
                        "mutually_exclusive": True,
                        "allowed_on": ["table", "column"],
                        "options": [
                            {"name": "Tier1", "fqn": "Tier.Tier1", "description": "Highest business criticality."},
                            {"name": "Tier2", "fqn": "Tier.Tier2", "description": "Important but not highest."},
                        ],
                    },
                ],
            },
            "connection": {
                "host": "localhost",
                "port": 5432,
                "database": "demo",
                "driver": "postgresql",
                "timezone": "UTC",
            },
            "data_quality_summary": {
                "total_findings": 1,
                "by_check": {"null_check": 1},
            },
            "source_system_context": {
                "contacts": [{"name": "Alex", "role": "Owner", "email": "alex@example.com"}],
                "delete_management_instruction": "Use soft deletes",
                "restrictions": [{"table_name": "customers", "type": "privacy", "scope": "internal", "details": "Mask exports", "owner": "security"}],
                "db_analysis_config": {
                    "exclude_schemas": ["audit"],
                    "exclude_tables": ["event_log", "audit.entries"],
                    "max_row_limit": 25,
                },
            },
            "tables": [
                {
                    "table": "customers",
                    "schema": "public",
                    "table_description": "Customer master",
                    "glossary_terms": ["Retail.Customer"],
                    "classification_tags": ["Tier.Tier1"],
                    "primary_keys": ["customer_id"],
                    "foreign_keys": [],
                    "row_count": 10,
                    "row_count_projection_1y": 15,
                    "row_count_projection_2y": 20,
                    "row_count_projection_5y": 35,
                    "field_classifications": {"email": "contact"},
                    "sensitive_fields": {"email": "pii_contact"},
                    "incremental_columns": ["updated_at"],
                    "partition_columns": ["created_at"],
                    "partition_columns_candidates": ["created_at", "updated_at"],
                    "join_candidates": [{"column": "sales_rep_id", "target_table": "employees", "target_column": "employee_id", "confidence": 0.88}],
                    "cdc_enabled": False,
                    "has_primary_key": True,
                    "has_foreign_keys": False,
                    "has_sensitive_fields": True,
                    "classification_summary": {"concept_counts": {"contact.email": 1}},
                    "unit_summary": {"columns_with_units": 0},
                    "columns": [
                        {
                            "name": "customer_id",
                            "type": "integer",
                            "nullable": False,
                            "is_incremental": False,
                            "cardinality": 10,
                            "null_count": 0,
                            "data_range": {"min": 1, "max": 10},
                            "data_category": "identifier",
                            "semantic_class": "customer_id",
                            "column_description": "Primary key",
                            "glossary_terms": ["Retail.Customer.Identifier"],
                            "classification_tags": [],
                            "concept_id": "identifier.customer",
                            "concept_confidence": 0.98,
                            "concept_evidence": {"rule": "name"},
                            "concept_sources": ["name"],
                            "concept_alias_group": "customer",
                            "unit_context": {},
                        },
                        {
                            "name": "email",
                            "type": "text",
                            "nullable": True,
                            "is_incremental": False,
                            "cardinality": 10,
                            "null_count": 1,
                            "data_range": {},
                            "data_category": "contact",
                            "semantic_class": "email",
                            "column_description": "Email address",
                            "glossary_terms": ["Retail.Customer.Email"],
                            "classification_tags": ["PII.Sensitive"],
                            "concept_id": "contact.email",
                            "concept_confidence": 0.91,
                            "concept_evidence": {"rule": "regex"},
                            "concept_sources": ["regex"],
                            "concept_alias_group": "contact",
                            "unit_context": {},
                        },
                    ],
                    "sample_data": {"email": ["a@example.com"]},
                    "data_quality": {
                        "findings": [
                            {
                                "check": "null_check",
                                "severity": "medium",
                                "column": "email",
                                "detail": "Null emails found",
                                "recommendation": "Backfill emails",
                            }
                        ]
                    },
                },
                {
                    "table": "employees",
                    "schema": "public",
                    "glossary_terms": [],
                    "classification_tags": ["Tier.Tier2"],
                    "primary_keys": ["employee_id"],
                    "foreign_keys": [{"column": "manager_id", "references": "employees.employee_id"}],
                    "row_count": 5,
                    "row_count_projection_1y": 5,
                    "row_count_projection_2y": 5,
                    "row_count_projection_5y": 5,
                    "field_classifications": {},
                    "sensitive_fields": {},
                    "incremental_columns": [],
                    "partition_columns": [],
                    "partition_columns_candidates": [],
                    "join_candidates": [],
                    "cdc_enabled": True,
                    "has_primary_key": True,
                    "has_foreign_keys": True,
                    "has_sensitive_fields": False,
                    "columns": [
                        {
                            "name": "employee_id",
                            "type": "integer",
                            "nullable": False,
                            "is_incremental": False,
                            "cardinality": 5,
                            "null_count": 0,
                            "data_range": {"min": 1, "max": 5},
                            "data_category": "identifier",
                            "semantic_class": "employee_id",
                            "column_description": "Employee key",
                            "glossary_terms": [],
                            "classification_tags": [],
                            "unit_context": {},
                        }
                    ],
                    "data_quality": {"findings": []},
                },
            ],
        }

    def _write_workbook(self, path):
        sheet_rows = json_to_excel._collect_sheets(self.payload)
        sheet_rows.update(json_to_excel._roundtrip_sheets(self.payload))
        json_to_excel._write_workbook(sheet_rows, path)

    def test_writer_uses_global_and_per_table_tabs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "schema.xlsx"
            self._write_workbook(output)

            wb = load_workbook(output, data_only=True)
            self.assertIn("Summary", wb.sheetnames)
            self.assertIn("SourceSystem", wb.sheetnames)
            self.assertIn("DataQualityFindings", wb.sheetnames)
            self.assertIn("Glossary", wb.sheetnames)
            self.assertIn("customers", wb.sheetnames)
            self.assertIn("employees", wb.sheetnames)
            self.assertNotIn("Tables", wb.sheetnames)
            self.assertNotIn("Columns", wb.sheetnames)
            self.assertNotIn("ForeignKeys", wb.sheetnames)
            self.assertNotIn("JoinCandidates", wb.sheetnames)

            customer_sections = _section_rows(wb["customers"])
            source_sections = _section_rows(wb["SourceSystem"])
            glossary_rows = list(wb["Glossary"].iter_rows(values_only=True))
            self.assertIn("Overview", customer_sections)
            self.assertIn("Columns", customer_sections)
            self.assertIn("DataQualityFindings", customer_sections)
            self.assertNotIn("FieldContextManual", source_sections)
            self.assertIn("DbAnalysisConfig", source_sections)
            self.assertIn("OpenMetadataClassifications", source_sections)
            self.assertNotIn("FieldClassifications", customer_sections)
            self.assertNotIn("SensitiveFields", customer_sections)
            self.assertEqual(glossary_rows[0], ("Field", "Description"))
            glossary_fields = {row[0] for row in glossary_rows[1:] if row and row[0]}
            self.assertTrue(
                {
                    "glossary_terms",
                    "om_class_<ClassificationName>",
                    "semantic_class",
                    "concept_id",
                    "concept_confidence",
                    "concept_alias_group",
                    "concept_evidence_json",
                    "concept_sources_json",
                    "unit",
                    "unit_source",
                    "canonical_unit",
                    "unit_system",
                    "unit_confidence",
                    "unit_notes",
                    "factor_to_canonical",
                    "offset_to_canonical",
                    "conversion_formula",
                }.issubset(glossary_fields)
            )
            glossary_descriptions = {row[0]: row[1] for row in glossary_rows[1:] if row and row[0]}
            self.assertEqual(glossary_descriptions["classification: PII"], "Personally identifiable information.")
            self.assertEqual(glossary_descriptions["classification option: PII.Sensitive"], "Sensitive personal data.")
            self.assertEqual(glossary_descriptions["classification: Tier"], "Business criticality tiers.")
            self.assertEqual(glossary_descriptions["classification option: Tier.Tier1"], "Highest business criticality.")
            self.assertEqual(list(source_sections["RestrictionsManual"][0].keys())[0], "table_name")
            self.assertEqual(source_sections["DbAnalysisConfig"][0]["exclude_schemas"], "audit")
            self.assertEqual(source_sections["DbAnalysisConfig"][0]["exclude_tables"], "event_log, audit.entries")
            self.assertEqual(source_sections["DbAnalysisConfig"][0]["max_row_limit"], 25)
            self.assertEqual(source_sections["OpenMetadataClassifications"][0]["classification_name"], "PII")
            self.assertEqual(source_sections["OpenMetadataClassifications"][0]["option_fqn"], "PII.Sensitive")
            self.assertIn("classification", customer_sections["Columns"][0])
            self.assertIn("sensitivity_label", customer_sections["Columns"][0])
            self.assertIn("om_class_PII", customer_sections["Columns"][0])
            self.assertIn("om_class_Tier", customer_sections["Columns"][0])
            self.assertNotIn("is_incremental", customer_sections["Columns"][0])
            self.assertNotIn("primary_keys", customer_sections["Overview"][0])
            self.assertNotIn("incremental_columns", customer_sections["Overview"][0])
            self.assertNotIn("partition_columns", customer_sections["Overview"][0])
            self.assertNotIn("partition_columns_candidates", customer_sections["Overview"][0])
            self.assertEqual(customer_sections["Overview"][0]["glossary_terms"], "Retail.Customer")
            self.assertNotIn("om_class_PII", customer_sections["Overview"][0])
            self.assertEqual(customer_sections["Overview"][0]["om_class_Tier"], "Tier.Tier1")
            self.assertEqual(customer_sections["Columns"][0]["glossary_terms"], "Retail.Customer.Identifier")
            self.assertEqual(customer_sections["Columns"][1]["glossary_terms"], "Retail.Customer.Email")
            self.assertEqual(customer_sections["Columns"][1]["om_class_PII"], "PII.Sensitive")
            self.assertIn(customer_sections["Columns"][1]["om_class_Tier"], ("", None))
            self.assertIn("row_count_projection_1y", customer_sections["Overview"][0])
            self.assertIn("row_count_projection_2y", customer_sections["Overview"][0])
            self.assertIn("row_count_projection_5y", customer_sections["Overview"][0])
            self.assertEqual(customer_sections["Overview"][0]["table_name"], "customers")
            self.assertEqual(customer_sections["Overview"][0]["row_count_projection_1y"], 15)
            self.assertEqual(customer_sections["Overview"][0]["row_count_projection_2y"], 20)
            self.assertEqual(customer_sections["Overview"][0]["row_count_projection_5y"], 35)

    def test_reader_applies_edits_from_source_and_table_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "schema.xlsx"
            self._write_workbook(output)

            wb = load_workbook(output)
            _set_section_value(wb["SourceSystem"], "Metadata", "value", "app", match_col="property", match_value="schema_filter")
            _set_section_value(wb["SourceSystem"], "DbAnalysisConfig", "exclude_schemas", "audit, staging")
            _set_section_value(wb["SourceSystem"], "DbAnalysisConfig", "exclude_tables", "event_log, temp_orders")
            _set_section_value(wb["SourceSystem"], "DbAnalysisConfig", "max_row_limit", 75)
            _set_section_value(wb["customers"], "Overview", "row_count", 25)
            _set_section_value(wb["customers"], "Overview", "row_count_projection_1y", 99)
            _set_section_value(wb["customers"], "Overview", "row_count_projection_2y", 199)
            _set_section_value(wb["customers"], "Overview", "row_count_projection_5y", 499)
            _set_section_value(wb["customers"], "Columns", "classification", "customer_contact", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "Columns", "sensitivity_label", "pii_email", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "Columns", "column_description", "Preferred email", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "Overview", "glossary_terms", "Retail.Customer, Retail.Buyer")
            _set_section_value(wb["customers"], "Overview", "om_class_Tier", "Tier.Tier2")
            _set_section_value(wb["customers"], "Columns", "glossary_terms", "Retail.Customer.PrimaryEmail", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "Columns", "om_class_PII", "PII.NonSensitive", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "Columns", "om_class_Tier", "Tier.Tier1", match_col="column_name", match_value="email")
            _set_section_value(wb["customers"], "DataQualityFindings", "detail", "Emails still missing", match_col="finding_index", match_value=1)
            wb["Glossary"]["B2"] = "Changed glossary text should be ignored"
            wb.save(output)

            edited_wb = load_workbook(output, data_only=True)
            payload = excel_to_json._read_roundtrip_payload(edited_wb)
            self.assertNotIn("Glossary", excel_to_json._visible_table_sheets(edited_wb))
            excel_to_json.apply_all_visible_edits(edited_wb, payload)

            self.assertEqual(payload["metadata"]["schema_filter"], "app")
            self.assertNotIn("field_context_manual", payload["source_system_context"])
            self.assertEqual(
                payload["source_system_context"]["db_analysis_config"],
                {
                    "exclude_schemas": ["audit", "staging"],
                    "exclude_tables": ["event_log", "temp_orders"],
                    "max_row_limit": 75,
                },
            )
            customers = next(table for table in payload["tables"] if table["table"] == "customers")
            self.assertEqual(customers["row_count"], 25)
            self.assertEqual(customers["row_count_projection_1y"], 15)
            self.assertEqual(customers["row_count_projection_2y"], 20)
            self.assertEqual(customers["row_count_projection_5y"], 35)
            self.assertEqual(customers["field_classifications"]["email"], "customer_contact")
            self.assertEqual(customers["sensitive_fields"]["email"], "pii_email")
            self.assertEqual(customers["glossary_terms"], ["Retail.Customer", "Retail.Buyer"])
            self.assertEqual(customers["classification_tags"], ["Tier.Tier2"])
            email_col = next(col for col in customers["columns"] if col["name"] == "email")
            self.assertEqual(email_col["column_description"], "Preferred email")
            self.assertEqual(email_col["glossary_terms"], ["Retail.Customer.PrimaryEmail"])
            self.assertEqual(sorted(email_col["classification_tags"]), ["PII.NonSensitive", "Tier.Tier1"])
            self.assertEqual(customers["data_quality"]["findings"][0]["detail"], "Emails still missing")

    def test_reader_rejects_invalid_classification_tags(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "schema.xlsx"
            self._write_workbook(output)

            wb = load_workbook(output)
            _set_section_value(wb["customers"], "Overview", "om_class_Tier", "Unknown.Tag")
            wb.save(output)

            edited_wb = load_workbook(output, data_only=True)
            payload = excel_to_json._read_roundtrip_payload(edited_wb)
            with self.assertRaisesRegex(ValueError, "Invalid classification tag 'Unknown.Tag' on table public.customers"):
                excel_to_json.apply_all_visible_edits(edited_wb, payload)

    def test_reader_rejects_table_level_pii_when_present_in_legacy_field(self):
        payload = json.loads(json.dumps(self.payload))
        payload["tables"][0]["classification_tags"] = ["PII.Sensitive"]
        with self.assertRaisesRegex(ValueError, "Classification 'PII' is not allowed on table public.customers"):
            excel_to_json._validate_classification_tags(
                payload,
                "PII.Sensitive",
                entity_level="table",
                entity_label="table public.customers",
            )

    def test_reader_rejects_mutually_exclusive_tags_in_legacy_field(self):
        payload = json.loads(json.dumps(self.payload))
        with self.assertRaisesRegex(ValueError, "mutually exclusive on column public.customers.email"):
            excel_to_json._validate_classification_tags(
                payload,
                "PII.Sensitive, PII.NonSensitive",
                entity_level="column",
                entity_label="column public.customers.email",
            )


if __name__ == "__main__":
    unittest.main()
