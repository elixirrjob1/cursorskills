import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path("/home/fillip/projec/cursorskills/.cursor/skills/schema-glossary-generator/scripts/generate_glossary.py")


def _load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


glossary_module = _load_module("generate_glossary", SCRIPT_PATH)


class SchemaGlossaryGeneratorTests(unittest.TestCase):
    def setUp(self):
        self.brief = {
            "domain": "Wholesale distribution",
            "description": "A company sells products to business customers, manages inventory, sends invoices, and collects payments.",
            "core_processes": ["order management", "procurement", "fulfillment"],
            "entities": ["customer", "supplier", "product"],
            "notes": "The business also handles deliveries and returns.",
        }

    def test_build_glossary_from_domain_brief(self):
        glossary = glossary_module.build_glossary(self.brief, "/tmp/domain.json")
        terms = {entry["term"]: entry for entry in glossary["entries"]}

        self.assertIn("Customer", terms)
        self.assertIn("Supplier", terms)
        self.assertIn("Product", terms)
        self.assertIn("Inventory", terms)
        self.assertIn("Invoice", terms)
        self.assertIn("Payment", terms)
        self.assertIn("Order Management", terms)

        customer = terms["Customer"]
        self.assertEqual(customer["term_type"], "business_entity")
        self.assertEqual(customer["status"], "draft")

        process = terms["Order Management"]
        self.assertEqual(process["term_type"], "business_process")

    def test_write_glossary_excel_creates_expected_tabs(self):
        glossary = glossary_module.build_glossary(self.brief, "/tmp/domain.json")
        with tempfile.TemporaryDirectory() as tmpdir:
            output = Path(tmpdir) / "domain_glossary.xlsx"
            glossary_module.write_glossary_excel(glossary, output)
            wb = load_workbook(output)
            self.assertEqual(wb.sheetnames, ["Glossary", "RunMetadata"])
            glossary_rows = list(wb["Glossary"].iter_rows(values_only=True))
            self.assertEqual(glossary_rows[0][0], "term")
            self.assertGreater(len(glossary_rows), 2)
            validations = list(wb["Glossary"].data_validations.dataValidation)
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].formula1, '"draft,approved,rejected"')
            self.assertIn("G2", str(validations[0].sqref))

    def test_output_paths_write_beside_input(self):
        input_path = Path("/tmp/wholesale-distribution.md")
        json_output, excel_output = glossary_module.output_paths(input_path)
        self.assertEqual(json_output.name, "wholesale-distribution_glossary.json")
        self.assertEqual(excel_output.name, "wholesale-distribution_glossary.xlsx")

    def test_run_writes_json_and_excel_outputs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "domain.json"
            input_path.write_text(json.dumps(self.brief), encoding="utf-8")
            json_output, excel_output, glossary = glossary_module.run(input_path)
            self.assertTrue(json_output.exists())
            self.assertTrue(excel_output.exists())
            written = json.loads(json_output.read_text(encoding="utf-8"))
            self.assertEqual(written["metadata"]["glossary_entry_count"], glossary["metadata"]["glossary_entry_count"])

    def test_schema_payload_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "schema.json"
            input_path.write_text(json.dumps({"tables": []}), encoding="utf-8")
            with self.assertRaises(ValueError):
                glossary_module.run(input_path)


if __name__ == "__main__":
    unittest.main()
